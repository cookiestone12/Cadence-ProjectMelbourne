"""
Import Jack Lomastro's complete placement catalog from his status sheet.
All songs in this file are placements where Jack is the producer/writer.
"""
import sys
import os
from datetime import datetime, date
import openpyxl

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from models.database import SessionLocal
from models.models import Creator, Song, Organization, SongCredit, ChecklistItem, SongChecklistStatus

def parse_percentage(value):
    """Parse percentage value from sheet (handles decimals like 0.24 = 24%).
    Always caps at 100% and rounds to 2 decimal places."""
    if value == '??' or value == 'N/A' or value is None:
        return None
    try:
        pct = float(value)
        # If value is less than 1, it's probably a decimal (0.24 = 24%)
        if pct < 1:
            pct = pct * 100
        # Cap at 100% and round to 2 decimal places
        pct = min(pct, 100.0)
        return round(pct, 2)
    except (ValueError, TypeError):
        return None

def parse_amount(value):
    """Parse dollar amount from sheet and convert to cents."""
    if value == '??' or value == 'N/A' or value is None:
        return None
    try:
        dollars = float(value)
        return int(dollars * 100)  # Convert to cents
    except (ValueError, TypeError):
        return None

def parse_yes_no(value):
    """Parse Yes/No values to boolean."""
    if value is None:
        return False
    value_str = str(value).strip().upper()
    return value_str in ['YES', 'Y', 'TRUE', '1', 'X', 'DONE', 'COMPLETED']

def parse_yes_no_na(value):
    """Parse Yes/No/N/A values as string."""
    if value is None:
        return 'N/A'
    value_str = str(value).strip().upper()
    if value_str in ['YES', 'Y', 'TRUE', '1', 'X', 'DONE', 'COMPLETED']:
        return 'Yes'
    elif value_str in ['NO', 'N', 'FALSE', '0']:
        return 'No'
    else:
        return 'N/A'

def parse_date(value):
    """Parse date from various formats."""
    if value is None:
        return None
    try:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            # Try common date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%m/%d/%y']:
                try:
                    return datetime.strptime(value.strip(), fmt).date()
                except:
                    continue
        return None
    except:
        return None

def import_jack_lomastro_catalog():
    """Import all of Jack Lomastro's placements from his placement status sheet."""
    
    # Load the Excel file
    excel_path = os.path.join(os.path.dirname(__file__), '..', 'attached_assets', 
                              'JACK LOMASTRO- PLACEMENT STATUS SHEET _1763865398551.xlsx')
    
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    db = SessionLocal()
    
    try:
        # Get Demo Label Co organization
        org = db.query(Organization).filter(Organization.name == "Demo Label Co.").first()
        if not org:
            print("Error: Demo Label Co. organization not found!")
            return
        
        print(f"Found organization: {org.name} (ID: {org.id})")
        
        # Create or get Jack Lomastro as a creator
        jack = db.query(Creator).filter(
            Creator.organization_id == org.id,
            Creator.display_name == "Jack Lomastro"
        ).first()
        
        if not jack:
            jack = Creator(
                organization_id=org.id,
                display_name="Jack Lomastro",
                legal_name="Jack Lomastro",
                roles=["PRODUCER", "SONGWRITER"],
                created_at=datetime.utcnow()
            )
            db.add(jack)
            db.commit()
            db.refresh(jack)
            print(f"Created creator: Jack Lomastro (ID: {jack.id})")
        else:
            print(f"Found existing creator: Jack Lomastro (ID: {jack.id})")
        
        # Get all checklist items for initializing new songs
        checklist_items = db.query(ChecklistItem).all()
        
        # Read and display headers
        headers = [cell.value for cell in ws[1]]
        print(f"\nExcel headers: {headers[:17]}")
        
        # Column mapping based on actual headers:
        # [0] ARTIST NAME → primary_artist
        # [1] SONG TITLE → title  
        # [2] PUBLISHING % → publishing_percentage
        # [3] ROYALTY → master_percentage (royalty rate)
        # [4] ADVANCE ($) → advance_amount (cents)
        # [5] LABEL → label
        # [6] CREDITED → has_contract_sent (credited/attributed)
        # [7] RECEIVED PAPERWORK → has_contract_executed (paperwork received)
        # [8] AGREEMENT → contract_location
        # [9] BMI REGISTRATION → is_registered_with_pro
        # [10] KOBALT REG → is_registered_with_dsp
        # [11] SOUNDEXCHANGE → soundexchange_registered
        # [12] PAYMENT RECEIVED → is_paid
        # [13] DATE RELEASED → release_date
        # [14] INVOICE SENT → is_invoiced
        # [15] NOTES → notes
        # [16] NOTES: 2 → notes (append)
        
        songs_created = 0
        songs_updated = 0
        credits_created = 0
        songs_skipped = 0
        
        print("\nImporting placements...")
        
        # Iterate through rows starting from row 2 (skip header)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if not row[0] and not row[1]:
                continue
            
            # Skip rows with no song title
            song_title = str(row[1]).strip() if row[1] else None
            if not song_title or song_title == 'None':
                songs_skipped += 1
                continue
            
            artist = str(row[0]).strip() if row[0] else "Unknown Artist"
            
            # Parse all fields
            publishing_pct = parse_percentage(row[2])
            master_pct = parse_percentage(row[3])
            advance = parse_amount(row[4])
            label = str(row[5]).strip() if row[5] else None
            
            # Contract/Credit status
            has_contract_sent = parse_yes_no(row[6])  # CREDITED
            has_contract_executed = parse_yes_no(row[7])  # RECEIVED PAPERWORK
            contract_location = str(row[8]).strip() if row[8] else None  # AGREEMENT
            
            # Registration status
            is_registered_with_pro = parse_yes_no(row[9])  # BMI REGISTRATION
            is_registered_with_dsp = parse_yes_no(row[10])  # KOBALT REG
            soundexchange_registered = parse_yes_no_na(row[11])  # SOUNDEXCHANGE
            
            # Payment status
            is_paid = parse_yes_no(row[12])  # PAYMENT RECEIVED
            release_date = parse_date(row[13])  # DATE RELEASED
            is_invoiced = parse_yes_no(row[14])  # INVOICE SENT
            
            # Notes
            notes_1 = str(row[15]).strip() if row[15] else ''
            notes_2 = str(row[16]).strip() if len(row) > 16 and row[16] else ''
            notes = f"{notes_1}\n{notes_2}".strip() if notes_1 or notes_2 else None
            
            # Determine payment status string
            if is_paid:
                payment_status = 'PAID'
            elif is_invoiced:
                payment_status = 'INVOICED'
            else:
                payment_status = 'PENDING'
            
            # Determine master_paid status
            master_paid = 'Yes' if is_paid else 'No'
            
            # Check if song already exists
            existing_song = db.query(Song).filter(
                Song.organization_id == org.id,
                Song.title == song_title,
                Song.primary_artist == artist
            ).first()
            
            if existing_song:
                # Update existing song with all placement data
                existing_song.publishing_percentage = publishing_pct
                existing_song.master_percentage = master_pct
                existing_song.advance_amount = advance
                existing_song.label = label
                existing_song.has_contract_sent = has_contract_sent
                existing_song.has_contract_executed = has_contract_executed
                existing_song.contract_location = contract_location
                existing_song.is_registered_with_pro = is_registered_with_pro
                existing_song.is_registered_with_dsp = is_registered_with_dsp
                existing_song.soundexchange_registered = soundexchange_registered
                existing_song.is_paid = is_paid
                existing_song.is_invoiced = is_invoiced
                existing_song.release_date = release_date
                existing_song.is_released = (release_date is not None)
                existing_song.payment_status = payment_status
                existing_song.master_paid = master_paid
                existing_song.notes = notes
                
                song = existing_song
                songs_updated += 1
            else:
                # Create new song
                song = Song(
                    organization_id=org.id,
                    title=song_title,
                    primary_artist=artist,
                    publishing_percentage=publishing_pct,
                    master_percentage=master_pct,
                    advance_amount=advance,
                    label=label,
                    has_contract_sent=has_contract_sent,
                    has_contract_executed=has_contract_executed,
                    contract_location=contract_location,
                    is_registered_with_pro=is_registered_with_pro,
                    is_registered_with_dsp=is_registered_with_dsp,
                    soundexchange_registered=soundexchange_registered,
                    is_paid=is_paid,
                    is_invoiced=is_invoiced,
                    release_date=release_date,
                    is_released=(release_date is not None),
                    payment_status=payment_status,
                    master_paid=master_paid,
                    notes=notes,
                    created_at=datetime.utcnow()
                )
                db.add(song)
                db.flush()  # Get the song ID
                
                # Create checklist items for new song
                for item in checklist_items:
                    status = SongChecklistStatus(
                        song_id=song.id,
                        checklist_item_id=item.id,
                        status="NOT_STARTED"
                    )
                    db.add(status)
                
                songs_created += 1
            
            # Check if Jack Lomastro already has credit on this song
            existing_credit = db.query(SongCredit).filter(
                SongCredit.song_id == song.id,
                SongCredit.creator_id == jack.id
            ).first()
            
            if not existing_credit:
                # Create credit linking song to Jack Lomastro as Producer
                credit = SongCredit(
                    song_id=song.id,
                    creator_id=jack.id,
                    role="Producer",
                    share_percentage=publishing_pct or 100.0  # Use publishing % as share
                )
                db.add(credit)
                credits_created += 1
            
            # Progress indicator every 100 rows
            if row_num % 100 == 0:
                print(f"  Processed {row_num} rows...")
        
        db.commit()
        
        # Get final count of Jack's songs
        jack_song_count = db.query(SongCredit).filter(
            SongCredit.creator_id == jack.id
        ).count()
        
        print(f"\n{'='*50}")
        print(f"Import Complete!")
        print(f"{'='*50}")
        print(f"Songs created: {songs_created}")
        print(f"Songs updated: {songs_updated}")
        print(f"Songs skipped: {songs_skipped}")
        print(f"Credits created: {credits_created}")
        print(f"Total placements for Jack Lomastro: {jack_song_count}")
        
    except Exception as e:
        db.rollback()
        print(f"Error during import: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

if __name__ == "__main__":
    import_jack_lomastro_catalog()
