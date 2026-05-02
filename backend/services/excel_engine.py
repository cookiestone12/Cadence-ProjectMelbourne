"""Unified branded Excel engine — single source of truth for every Cadence XLSX.

Routes call :class:`BrandedWorkbook`, add sheets via :meth:`add_sheet`, and
get the bytes via :meth:`build`. Every workbook starts with a title block
("display_name" / report title / "Generated …" / "Powered by Cadence")
in the org's primary color, and each sheet uses a branded header row
(primary color background, white bold text), zebra striping in a tinted
shade of the primary color, and column auto-width.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Any, Iterable, List, Optional, Sequence, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .branding import OrgTheme, POWERED_BY, lighten_hex, parse_hex_color, theme_from_org

logger = logging.getLogger("cadence")

__all__ = [
    "BrandedWorkbook",
    "OrgTheme",
    "theme_from_org",
]


def _argb(hex_color: str) -> str:
    """Convert ``#RRGGBB`` to openpyxl's ``AARRGGBB`` (opaque) form."""
    h = parse_hex_color(hex_color)
    return "FF" + h[1:].upper()


_THIN_SIDE = Side(border_style="thin", color="FFD0D5D1")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


class BrandedWorkbook:
    """Branded Excel workbook builder.

    Usage::

        wb = BrandedWorkbook(theme, title="Schedule A — Q3 2025")
        wb.add_sheet(
            name="Released",
            headers=["Title", "Artist", "ISRC", "Status"],
            rows=[...],
        )
        wb.add_sheet(name="Pipeline", headers=[...], rows=[...])
        bytes_ = wb.build()
    """

    TITLE_BLOCK_ROWS = 5  # rows reserved at the top of every sheet for the title block

    def __init__(
        self,
        theme: OrgTheme,
        title: str,
        subtitle: Optional[str] = None,
        include_title_block: bool = True,
    ) -> None:
        self.theme = theme
        self.title = title
        self.subtitle = subtitle
        self.include_title_block = include_title_block
        self._wb = Workbook()
        # Remove the default sheet — sheets are added explicitly
        default = self._wb.active
        self._wb.remove(default)
        self._sheet_count = 0

    # ------------------------------------------------------------------
    # Sheet construction
    # ------------------------------------------------------------------
    def add_sheet(
        self,
        name: str,
        headers: Sequence[str],
        rows: Iterable[Sequence[Any]],
        *,
        column_widths: Optional[Sequence[Optional[float]]] = None,
        freeze_header: bool = True,
        zebra: bool = True,
        section_subtitle: Optional[str] = None,
    ) -> Worksheet:
        """Add a sheet with branded title block, header row, and rows.

        - ``headers`` becomes a primary-colored header row.
        - ``rows`` is iterated lazily; each cell is written verbatim.
        - ``column_widths`` lets a caller pin specific widths; ``None`` entries
          fall back to auto-width (clamped to a sane min/max).
        - ``zebra`` toggles per-row alternating backgrounds.
        - ``section_subtitle`` is rendered above the header row inside the
          title block (e.g. "Released catalog — 42 entries").
        """
        ws = self._wb.create_sheet(self._unique_sheet_name(name))
        primary_argb = _argb(self.theme.primary_color)
        zebra_argb = _argb(self.theme.zebra_color)
        ink_argb = _argb(self.theme.text_color)
        muted_argb = _argb(self.theme.muted_color)

        start_row = 1
        if self.include_title_block:
            self._write_title_block(
                ws,
                primary_argb=primary_argb,
                ink_argb=ink_argb,
                muted_argb=muted_argb,
                num_columns=max(1, len(headers)),
                section_subtitle=section_subtitle,
            )
            start_row = self.TITLE_BLOCK_ROWS + 1

        # Header row
        header_fill = PatternFill("solid", fgColor=primary_argb)
        header_font = Font(name="Helvetica", bold=True, size=10, color="FFFFFFFF")
        header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(label))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = _BORDER
        ws.row_dimensions[start_row].height = 22

        # Data rows
        body_font = Font(name="Helvetica", size=10, color=ink_argb)
        body_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
        body_fill = PatternFill("solid", fgColor=zebra_argb) if zebra else None

        data_start = start_row + 1
        max_lengths = [len(str(h)) for h in headers]
        row_count = 0
        for r_idx, row in enumerate(rows):
            target_row = data_start + r_idx
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=target_row, column=c_idx, value=value)
                cell.font = body_font
                cell.alignment = body_align
                cell.border = _BORDER
                if zebra and r_idx % 2 == 1 and body_fill is not None:
                    cell.fill = body_fill
                if c_idx - 1 < len(max_lengths):
                    text = "" if value is None else str(value)
                    if len(text) > max_lengths[c_idx - 1]:
                        max_lengths[c_idx - 1] = len(text)
            row_count = r_idx + 1

        # Column widths
        if column_widths is not None:
            for c_idx, width in enumerate(column_widths, start=1):
                if width is not None:
                    ws.column_dimensions[get_column_letter(c_idx)].width = float(width)
                else:
                    self._auto_width(ws, c_idx, max_lengths)
        else:
            for c_idx in range(1, len(headers) + 1):
                self._auto_width(ws, c_idx, max_lengths)

        if freeze_header:
            ws.freeze_panes = ws.cell(row=data_start, column=1)

        self._sheet_count += 1
        return ws

    def add_kpi_sheet(
        self,
        name: str,
        kpis: Sequence[dict],
        *,
        section_subtitle: Optional[str] = None,
    ) -> Worksheet:
        """Render a vertical KPI table (Metric / Value)."""
        rows = [(k.get("label", ""), k.get("value", "")) for k in kpis]
        return self.add_sheet(
            name=name,
            headers=["Metric", "Value"],
            rows=rows,
            column_widths=[36, 30],
            zebra=True,
            section_subtitle=section_subtitle,
        )

    def add_section_sheet(
        self,
        name: str,
        sections: Sequence[dict],
    ) -> Worksheet:
        """Render multiple labeled (headers, rows) sections in a single sheet.

        Each section is a dict ``{"title": str, "headers": [...], "rows": [...]}``.
        Sections stack vertically with the section title in the org primary color
        and a blank row separator in between.
        """
        if not sections:
            return self.add_sheet(name, headers=["(no data)"], rows=[])
        # Build the first section as the base sheet, then append the rest by
        # writing into the same worksheet directly so they share the title block.
        first = sections[0]
        ws = self.add_sheet(
            name=name,
            headers=first.get("headers", []),
            rows=first.get("rows", []),
            section_subtitle=first.get("title"),
        )
        primary_argb = _argb(self.theme.primary_color)
        ink_argb = _argb(self.theme.text_color)
        cur_row = ws.max_row + 2

        for section in sections[1:]:
            headers = section.get("headers", [])
            rows = list(section.get("rows", []))
            title = section.get("title")
            if title:
                cell = ws.cell(row=cur_row, column=1, value=title)
                cell.font = Font(name="Helvetica", bold=True, size=12, color=primary_argb)
                cell.alignment = Alignment(horizontal="left")
                ws.merge_cells(
                    start_row=cur_row, start_column=1,
                    end_row=cur_row, end_column=max(1, len(headers)),
                )
                cur_row += 1
            # Header row
            for c_idx, label in enumerate(headers, start=1):
                cell = ws.cell(row=cur_row, column=c_idx, value=str(label))
                cell.fill = PatternFill("solid", fgColor=primary_argb)
                cell.font = Font(name="Helvetica", bold=True, size=10, color="FFFFFFFF")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = _BORDER
            cur_row += 1
            for row in rows:
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=cur_row, column=c_idx, value=value)
                    cell.font = Font(name="Helvetica", size=10, color=ink_argb)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = _BORDER
                cur_row += 1
            cur_row += 1  # spacer
        return ws

    # ------------------------------------------------------------------
    # Internals
    # ------------------------------------------------------------------
    def _unique_sheet_name(self, name: str) -> str:
        """Excel limits sheet names to 31 chars + must be unique."""
        clean = (name or "Sheet").replace("/", "-").replace("\\", "-")[:31]
        existing = {s.title for s in self._wb.worksheets}
        if clean not in existing:
            return clean
        for i in range(2, 100):
            candidate = (clean[:28] + f"_{i}")[:31]
            if candidate not in existing:
                return candidate
        return clean[:28] + "_X"

    def _write_title_block(
        self,
        ws: Worksheet,
        *,
        primary_argb: str,
        ink_argb: str,
        muted_argb: str,
        num_columns: int,
        section_subtitle: Optional[str] = None,
    ) -> None:
        """Render the 5-row title block at the top of every branded sheet."""
        last_col = max(1, num_columns)

        # Row 1: org display name
        c1 = ws.cell(row=1, column=1, value=self.theme.display_name or self.theme.name)
        c1.font = Font(name="Helvetica", bold=True, size=14, color=ink_argb)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws.row_dimensions[1].height = 20

        # Row 2: report title in primary color
        c2 = ws.cell(row=2, column=1, value=self.title)
        c2.font = Font(name="Helvetica", bold=True, size=18, color=primary_argb)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        ws.row_dimensions[2].height = 26

        # Row 3: subtitle / section subtitle / generated stamp
        sub_parts = []
        if self.subtitle:
            sub_parts.append(self.subtitle)
        if section_subtitle:
            sub_parts.append(section_subtitle)
        sub_parts.append(f"Generated {datetime.utcnow().strftime('%B %d, %Y %H:%M UTC')}")
        c3 = ws.cell(row=3, column=1, value=" — ".join(sub_parts))
        c3.font = Font(name="Helvetica", italic=True, size=10, color=muted_argb)
        c3.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

        # Row 4: small "Powered by Cadence" caption
        c4 = ws.cell(row=4, column=1, value=self.theme.powered_by_text)
        c4.font = Font(name="Helvetica", bold=True, size=8, color=primary_argb)
        c4.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)

        # Row 5: branded accent line via filled cells in primary color
        for col in range(1, last_col + 1):
            cell = ws.cell(row=5, column=col, value="")
            cell.fill = PatternFill("solid", fgColor=primary_argb)
        ws.row_dimensions[5].height = 4

    def _auto_width(self, ws: Worksheet, col_idx: int, max_lengths: list) -> None:
        try:
            length = max_lengths[col_idx - 1]
        except IndexError:
            length = 12
        # Add a little padding, clamp to sensible bounds
        width = max(10.0, min(60.0, float(length) + 4.0))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ------------------------------------------------------------------
    # Output
    # ------------------------------------------------------------------
    def build(self) -> bytes:
        """Save the workbook and return the bytes."""
        if not self._wb.worksheets:
            # openpyxl refuses to save an empty workbook
            self.add_sheet("Empty", headers=["(no data)"], rows=[])
        buf = io.BytesIO()
        self._wb.save(buf)
        return buf.getvalue()


# ---------------------------------------------------------------------------
# Convenience helpers used by route handlers
# ---------------------------------------------------------------------------

def excel_response_headers(filename: str) -> dict:
    """Return the headers for a downloadable XLSX response."""
    return {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


def pdf_response_headers(filename: str) -> dict:
    return {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/pdf",
    }


def csv_response_headers(filename: str) -> dict:
    return {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "text/csv",
    }
