from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, date, timedelta
from ..models import get_db
from ..models.models import (
    User, Organization, OrganizationMember, Song, SongCredit,
    ValuationCalculation, SongStreamingMetrics, TerritoryRevenue,
    Creator, UnderwritingRun, RoyaltyStatement
)
from ..utils.auth import get_current_user
from ..services.underwriting_engine import run_underwriting
from ..services.underwriting_controls import run_reconciliation_controls
from ..services.valuation_engine import (
    compute_source_typed_valuation,
    compute_full_catalog_valuation,
)
import io
import json

router = APIRouter(prefix="/api/valuation", tags=["Valuations"])

# Org-scoped catalog valuation router (Task #172 spec contract).
# Mirrored to /api/v1/organizations/{org_id}/valuation/catalog by
# main._mount_v1_routes() the same way the rest of the API is.
org_router = APIRouter(prefix="/api/organizations", tags=["Valuations"])

class SongValuationDetail(BaseModel):
    song_id: int
    title: str
    primary_artist: str
    isrc: Optional[str]
    release_date: Optional[str]
    
    total_streams: int
    ownership_percentage: float
    
    streaming_multiple_value: float
    revenue_multiple_value: float
    market_comp_value: float
    black_box_value: float
    final_valuation: float
    
    thirty_day_revenue: float
    ninety_day_revenue: float
    annual_revenue: float
    
    growth_rate: float
    risk_score: float
    
    class Config:
        from_attributes = True

class TerritoryRevenueDetail(BaseModel):
    territory_code: str
    territory_name: str
    total_streams: int
    publishing_revenue: float
    master_revenue: float
    total_revenue: float
    
    class Config:
        from_attributes = True

class CatalogValuationSummary(BaseModel):
    organization_name: str
    total_songs: int
    total_catalog_value: float
    total_thirty_day_revenue: float
    total_annual_revenue: float
    avg_growth_rate: float
    generated_at: str
    
    top_songs: List[SongValuationDetail]
    territory_breakdown: List[TerritoryRevenueDetail]

class SongDetailWithValuation(BaseModel):
    song: Dict[str, Any]
    streaming_metrics: Dict[str, Any]
    territory_revenues: List[Dict[str, Any]]
    valuation: Dict[str, Any]
    credits: List[Dict[str, Any]]

@router.get("/catalog/summary", summary="Get comprehensive catalog valuation summary for the organization", description="Returns the comprehensive catalog valuation summary used by the valuation dashboard (NPV band, multiples, top contributors, last underwriting run).\n\n**Query:** `org_id` (defaults to caller's current org), `discount_rate?`, `multiple?`, `scope_creator_id?` (restricts the summary to songs credited to a single creator in the caller's org).\n**Auth:** Bearer JWT — caller must be a member of the org.\n**Response:** `{ npv_cents, low_high_band: [low, high], multiple_value_cents, top_contributors: [...], last_run_at }`.")
def get_catalog_valuation_summary(
    scope_creator_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get comprehensive catalog valuation summary for the organization.

    When ``scope_creator_id`` is provided, the summary is restricted to songs
    credited to that creator. The creator must belong to the caller's org or
    the request is rejected with 404 to avoid cross-org data exposure.
    """
    
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")
    
    org = db.query(Organization).filter(Organization.id == membership.organization_id).first()

    scoped_song_ids: Optional[set] = None
    if scope_creator_id is not None:
        from ..models.models import Creator as _Creator, SongCredit as _SongCredit
        creator_in_org = db.query(_Creator.id).filter(
            _Creator.id == scope_creator_id,
            _Creator.organization_id == org.id,
        ).first()
        if not creator_in_org:
            raise HTTPException(status_code=404, detail="Creator not found in this org")
        scoped_song_ids = {
            sid for (sid,) in db.query(_SongCredit.song_id)
            .filter(_SongCredit.creator_id == scope_creator_id)
            .distinct()
            .all()
        }
        if scoped_song_ids:
            songs = db.query(Song).filter(
                Song.organization_id == org.id,
                Song.id.in_(scoped_song_ids),
            ).all()
        else:
            songs = []
    else:
        songs = db.query(Song).filter(Song.organization_id == org.id).all()
    
    valuations_q = db.query(ValuationCalculation).filter(
        ValuationCalculation.organization_id == org.id
    )
    if scoped_song_ids is not None:
        # Restrict org-wide valuation rows to the scoped creator's songs so
        # totals/top-songs/avg-growth reflect ONLY that creator's catalog.
        if scoped_song_ids:
            valuations_q = valuations_q.filter(
                ValuationCalculation.song_id.in_(scoped_song_ids)
            )
        else:
            valuations_q = valuations_q.filter(False)
    valuations = valuations_q.order_by(ValuationCalculation.calculation_date.desc()).all()
    
    valuation_by_song = {}
    for v in valuations:
        if v.song_id not in valuation_by_song:
            valuation_by_song[v.song_id] = v
    
    total_catalog_value = sum(v.final_valuation_cents for v in valuation_by_song.values()) / 100 if valuation_by_song else 0.0
    total_thirty_day_revenue = sum(v.thirty_day_revenue_cents for v in valuation_by_song.values()) / 100 if valuation_by_song else 0.0
    total_annual_revenue = sum(v.annual_revenue_cents for v in valuation_by_song.values()) / 100 if valuation_by_song else 0.0
    
    growth_rates = [v.growth_rate for v in valuation_by_song.values() if v.growth_rate]
    avg_growth_rate = sum(growth_rates) / len(growth_rates) if growth_rates else 0.0
    
    top_valuations = sorted(valuation_by_song.values(), key=lambda v: v.final_valuation_cents, reverse=True)[:10] if valuation_by_song else []
    
    top_songs = []
    for val in top_valuations:
        song = next((s for s in songs if s.id == val.song_id), None)
        if song:
            metrics = db.query(SongStreamingMetrics).filter(
                SongStreamingMetrics.song_id == song.id
            ).order_by(SongStreamingMetrics.period_date.desc()).first()
            
            top_songs.append({
                "song_id": song.id,
                "title": song.title,
                "primary_artist": song.primary_artist,
                "isrc": song.isrc,
                "release_date": song.release_date.isoformat() if song.release_date else None,
                "total_streams": metrics.total_streams if metrics else 0,
                "ownership_percentage": metrics.ownership_percentage if metrics else 1.0,
                "streaming_multiple_value": val.streaming_multiple_value_cents / 100,
                "revenue_multiple_value": val.revenue_multiple_value_cents / 100,
                "market_comp_value": val.market_comp_value_cents / 100,
                "black_box_value": val.black_box_value_cents / 100,
                "final_valuation": val.final_valuation_cents / 100,
                "thirty_day_revenue": val.thirty_day_revenue_cents / 100,
                "ninety_day_revenue": val.ninety_day_revenue_cents / 100,
                "annual_revenue": val.annual_revenue_cents / 100,
                "growth_rate": val.growth_rate,
                "risk_score": val.risk_score
            })
    
    territory_q = db.query(
        TerritoryRevenue.territory_code,
        TerritoryRevenue.territory_name,
        func.sum(TerritoryRevenue.total_streams).label("total_streams"),
        func.sum(TerritoryRevenue.publishing_revenue_cents).label("publishing_revenue_cents"),
        func.sum(TerritoryRevenue.master_revenue_cents).label("master_revenue_cents"),
        func.sum(TerritoryRevenue.total_revenue_cents).label("total_revenue_cents")
    ).filter(
        TerritoryRevenue.organization_id == org.id
    )
    if scoped_song_ids is not None:
        if scoped_song_ids:
            territory_q = territory_q.filter(TerritoryRevenue.song_id.in_(scoped_song_ids))
        else:
            territory_q = territory_q.filter(False)
    territory_data = territory_q.group_by(
        TerritoryRevenue.territory_code,
        TerritoryRevenue.territory_name
    ).order_by(desc("total_revenue_cents")).all()
    
    territory_breakdown = [
        {
            "territory_code": t.territory_code,
            "territory_name": t.territory_name,
            "total_streams": t.total_streams,
            "publishing_revenue": t.publishing_revenue_cents / 100,
            "master_revenue": t.master_revenue_cents / 100,
            "total_revenue": t.total_revenue_cents / 100
        }
        for t in territory_data
    ]
    
    return {
        "organization_name": org.display_name or org.name,
        "total_songs": len(songs),
        "total_catalog_value": total_catalog_value,
        "total_thirty_day_revenue": total_thirty_day_revenue,
        "total_annual_revenue": total_annual_revenue,
        "avg_growth_rate": avg_growth_rate,
        "generated_at": datetime.utcnow().isoformat(),
        "top_songs": top_songs,
        "territory_breakdown": territory_breakdown
    }

@router.get("/song/{song_id}/detail", summary="Get detailed valuation information for a specific song", description="Returns detailed valuation information for a specific song (cashflow history, decay fit, NPV breakdown).\n\n**Path parameter:** `song_id`.\n**Auth:** Bearer JWT — caller must be a member of the song's org.\n**Response:** `{ song_id, title, npv_cents, historical: [{period, cents}], projected: [{period, cents}], decay: {a, k, r2}, share_of_catalog_pct }`.")
def get_song_valuation_detail(
    song_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get detailed valuation information for a specific song"""
    
    song = db.query(Song).filter(Song.id == song_id).first()
    if not song:
        raise HTTPException(status_code=404, detail="Song not found")
    
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id,
        OrganizationMember.organization_id == song.organization_id
    ).first()
    
    if not membership:
        raise HTTPException(status_code=403, detail="Not authorized to access this song")
    
    valuation = db.query(ValuationCalculation).filter(
        ValuationCalculation.song_id == song_id
    ).order_by(ValuationCalculation.calculation_date.desc()).first()
    
    metrics = db.query(SongStreamingMetrics).filter(
        SongStreamingMetrics.song_id == song_id
    ).order_by(SongStreamingMetrics.period_date.desc()).first()
    
    territory_revenues = db.query(TerritoryRevenue).filter(
        TerritoryRevenue.song_id == song_id
    ).order_by(TerritoryRevenue.total_revenue_cents.desc()).all()
    
    credits = db.query(SongCredit, Creator).join(
        Creator, SongCredit.creator_id == Creator.id
    ).filter(SongCredit.song_id == song_id).all()
    
    return {
        "song": {
            "id": song.id,
            "title": song.title,
            "primary_artist": song.primary_artist,
            "isrc": song.isrc,
            "iswc": song.iswc,
            "release_date": song.release_date.isoformat() if song.release_date else None,
            "health_score": song.status_health_score
        },
        "streaming_metrics": {
            "total_streams": metrics.total_streams if metrics else 0,
            "ad_supported_streams": metrics.ad_supported_streams if metrics else 0,
            "premium_streams": metrics.premium_streams if metrics else 0,
            "interactive_streams": metrics.interactive_streams if metrics else 0,
            "on_demand_streams": metrics.on_demand_streams if metrics else 0,
            "programmed_streams": metrics.programmed_streams if metrics else 0,
            "audio_streams": metrics.audio_streams if metrics else 0,
            "video_streams": metrics.video_streams if metrics else 0,
            "song_sales": metrics.song_sales if metrics else 0,
            "ownership_percentage": metrics.ownership_percentage if metrics else 1.0
        } if metrics else {},
        "territory_revenues": [
            {
                "territory_code": tr.territory_code,
                "territory_name": tr.territory_name,
                "total_streams": tr.total_streams,
                "publishing_revenue": tr.publishing_revenue_cents / 100,
                "master_revenue": tr.master_revenue_cents / 100,
                "total_revenue": tr.total_revenue_cents / 100
            }
            for tr in territory_revenues
        ],
        "valuation": {
            "streaming_multiple_value": valuation.streaming_multiple_value_cents / 100 if valuation else 0,
            "revenue_multiple_value": valuation.revenue_multiple_value_cents / 100 if valuation else 0,
            "market_comp_value": valuation.market_comp_value_cents / 100 if valuation else 0,
            "black_box_value": valuation.black_box_value_cents / 100 if valuation else 0,
            "final_valuation": valuation.final_valuation_cents / 100 if valuation else 0,
            "methodology": valuation.valuation_methodology if valuation else "N/A",
            "thirty_day_revenue": valuation.thirty_day_revenue_cents / 100 if valuation else 0,
            "ninety_day_revenue": valuation.ninety_day_revenue_cents / 100 if valuation else 0,
            "annual_revenue": valuation.annual_revenue_cents / 100 if valuation else 0,
            "growth_rate": valuation.growth_rate if valuation else 0,
            "risk_score": valuation.risk_score if valuation else 0.5
        } if valuation else {},
        "credits": [
            {
                "creator_id": credit.SongCredit.creator_id,
                "creator_name": credit.Creator.display_name,
                "role": credit.SongCredit.role,
                "share_percentage": credit.SongCredit.share_percentage
            }
            for credit in credits
        ]
    }

@router.get("/catalog/download/excel", summary="Download catalog valuation report as Excel file", description='Renders the catalog valuation report as an Excel workbook for download.\n\n**Query:** `org_id`, `discount_rate?`, `multiple?`.\n**Auth:** Bearer JWT — caller must be a member of the org.\n**Response:** `application/vnd.openxmlformats-officedocument.spreadsheetml.sheet` download.')
def download_catalog_valuation_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download catalog valuation report as Excel file"""
    
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")
    
    org = db.query(Organization).filter(Organization.id == membership.organization_id).first()
    
    wb = Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "Catalog Summary"
    
    ws_summary['A1'] = "Cadence Catalog Report"
    ws_summary['A1'].font = Font(size=14, bold=True)
    ws_summary['A2'] = f"Catalog: {org.name}"
    ws_summary['A3'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    
    ws_summary['A5'] = "Metric"
    ws_summary['B5'] = "Value"
    ws_summary['A5'].font = Font(bold=True)
    ws_summary['B5'].font = Font(bold=True)
    
    songs = db.query(Song).filter(Song.organization_id == org.id).all()
    valuations = db.query(ValuationCalculation).filter(
        ValuationCalculation.organization_id == org.id
    ).all()
    
    valuation_by_song = {}
    for v in valuations:
        if v.song_id not in valuation_by_song:
            valuation_by_song[v.song_id] = v
    
    total_catalog_value = sum(v.final_valuation_cents for v in valuation_by_song.values()) / 100
    total_annual_revenue = sum(v.annual_revenue_cents for v in valuation_by_song.values()) / 100
    
    ws_summary['A6'] = "Total Songs"
    ws_summary['B6'] = len(songs)
    ws_summary['A7'] = "Total Catalog Value"
    ws_summary['B7'] = f"${total_catalog_value:,.2f}"
    ws_summary['A8'] = "Annual Revenue"
    ws_summary['B8'] = f"${total_annual_revenue:,.2f}"
    
    ws_details = wb.create_sheet("Song Details")
    headers = ['Title', 'Artist', 'ISRC', 'Release Date', 'Total Streams', 'Valuation', 'Annual Revenue', 'Growth Rate']
    for col, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    
    row = 2
    for song in songs:
        val = valuation_by_song.get(song.id)
        metrics = db.query(SongStreamingMetrics).filter(
            SongStreamingMetrics.song_id == song.id
        ).order_by(SongStreamingMetrics.period_date.desc()).first()
        
        ws_details.cell(row=row, column=1, value=song.title)
        ws_details.cell(row=row, column=2, value=song.primary_artist)
        ws_details.cell(row=row, column=3, value=song.isrc or "")
        ws_details.cell(row=row, column=4, value=song.release_date.isoformat() if song.release_date else "")
        ws_details.cell(row=row, column=5, value=metrics.total_streams if metrics else 0)
        ws_details.cell(row=row, column=6, value=f"${val.final_valuation_cents / 100:,.2f}" if val else "$0.00")
        ws_details.cell(row=row, column=7, value=f"${val.annual_revenue_cents / 100:,.2f}" if val else "$0.00")
        ws_details.cell(row=row, column=8, value=f"{val.growth_rate * 100:.1f}%" if val else "0.0%")
        row += 1
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"cadence_catalog_report_{org.name.replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


class UnderwritingRunRequest(BaseModel):
    periodization_mode: str = "activity"
    granularity: str = "half"
    exclude_right_types: List[str] = []
    exclude_flags: List[str] = []
    scope_creator_id: Optional[int] = None
    include_sync: bool = True
    use_gross: bool = False


@router.post(
    "/underwriting/run",
    summary='Trigger a fresh underwriting run on the catalog',
    description="Kicks off the underwriting engine which produces a snapshot of the catalog's value: spine groupings, decay fits, concentration metrics. Synchronous; returns when complete.\n\n**Body:** `{ org_id: int, params?: {discount_rate, horizon_years, ...} }`.\n**Auth:** Bearer JWT — caller must be a member of the org.\n**Response:** `{ run_id, status, started_at, completed_at }`.",
)
def trigger_underwriting_run(
    request: UnderwritingRunRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    if request.scope_creator_id is not None:
        from ..models.models import Creator as _Creator
        belongs = db.query(_Creator.id).filter(
            _Creator.id == request.scope_creator_id,
            _Creator.organization_id == membership.organization_id,
        ).first()
        if not belongs:
            raise HTTPException(status_code=404, detail="Creator not found in this org")

    try:
        result = run_underwriting(
            db=db,
            org_id=membership.organization_id,
            user_id=current_user.id,
            periodization_mode=request.periodization_mode,
            granularity=request.granularity,
            exclude_right_types=request.exclude_right_types or None,
            exclude_flags=request.exclude_flags or None,
            scope_creator_id=request.scope_creator_id,
            include_sync=request.include_sync,
            use_gross=request.use_gross,
        )
        db.commit()
        return result
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Underwriting run failed: {str(e)}")


@router.get(
    "/underwriting/runs",
    summary='List historical underwriting runs',
    description='Returns the audit trail of `/run` invocations.\n\n**Query:** `org_id`, `limit`, `offset`.\n**Auth:** Bearer JWT — caller must be a member of the org.\n**Response:** `{ runs: [{id, started_at, completed_at, params, npv_cents, status}] }`.',
)
def list_underwriting_runs(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    runs = db.query(UnderwritingRun).filter(
        UnderwritingRun.organization_id == membership.organization_id
    ).order_by(UnderwritingRun.created_at.desc()).limit(20).all()

    creator_ids = {r.scope_creator_id for r in runs if r.scope_creator_id}
    name_by_id: dict[int, str] = {}
    if creator_ids:
        from ..models.models import Creator as _Creator
        # Restrict creator-name lookup to the caller's org so a stray
        # cross-org scope_creator_id (legacy or malicious) cannot leak the
        # display name of a creator from another tenant.
        for c in db.query(_Creator.id, _Creator.display_name).filter(
            _Creator.id.in_(creator_ids),
            _Creator.organization_id == membership.organization_id,
        ).all():
            name_by_id[c.id] = c.display_name

    return [
        {
            "id": r.id,
            "status": r.status,
            "kb_version": r.kb_version,
            "inputs": r.inputs,
            "scope_creator_id": r.scope_creator_id,
            "scope_creator_name": name_by_id.get(r.scope_creator_id) if r.scope_creator_id else None,
            "portfolio_summary": r.outputs,
            "valuation": r.valuation_data.get("blended") if r.valuation_data else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None,
        }
        for r in runs
    ]


@router.get(
    "/underwriting/runs/{run_id}",
    summary='Get the headline result of a single underwriting run',
    description="Returns the run's summary numbers: NPV, multiples, top contributors, model parameters.\n\n**Path parameter:** `run_id`.\n**Auth:** Bearer JWT — caller must be a member of the run's org.\n**Response:** `{ id, params, summary: {...}, top_contributors: [...], completed_at }`.",
)
def get_underwriting_run(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    run = db.query(UnderwritingRun).filter(
        UnderwritingRun.id == run_id,
        UnderwritingRun.organization_id == membership.organization_id,
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Underwriting run not found")

    return {
        "id": run.id,
        "status": run.status,
        "kb_version": run.kb_version,
        "inputs": run.inputs,
        "portfolio_summary": run.outputs,
        "spine": run.spine_data,
        "decay": run.decay_data,
        "concentration": run.concentration_data,
        "projections": run.projection_data,
        "valuation": run.valuation_data,
        "exceptions": run.exceptions,
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "completed_at": run.completed_at.isoformat() if run.completed_at else None,
    }


@router.get(
    "/underwriting/runs/{run_id}/spine",
    summary='Get the spine breakdown of an underwriting run',
    description='The "spine" is the cohort grouping used by the model — returns per-cohort cashflows, weights, and projected NPV.\n\n**Path parameter:** `run_id`.\n**Auth:** Bearer JWT — caller must be a member of the run\'s org.\n**Response:** `{ spine: [{cohort_label, weight, historical: [...], projected: [...], npv_cents}] }`.',
)
def get_underwriting_spine(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    run = db.query(UnderwritingRun).filter(
        UnderwritingRun.id == run_id,
        UnderwritingRun.organization_id == membership.organization_id,
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Underwriting run not found")

    return run.spine_data or {"entries": [], "total_entries": 0}


@router.get(
    "/underwriting/runs/{run_id}/decay",
    summary='Get the decay-curve fits from an underwriting run',
    description="Returns the per-cohort exponential decay parameters used to project earnings.\n\n**Path parameter:** `run_id`.\n**Auth:** Bearer JWT — caller must be a member of the run's org.\n**Response:** `{ fits: [{cohort_label, a, k, r2, data_points}] }`.",
)
def get_underwriting_decay(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    run = db.query(UnderwritingRun).filter(
        UnderwritingRun.id == run_id,
        UnderwritingRun.organization_id == membership.organization_id,
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Underwriting run not found")

    return run.decay_data or {}


@router.get(
    "/underwriting/runs/{run_id}/concentration",
    summary='Get the concentration risk metrics for an underwriting run',
    description="How concentrated catalog earnings are in a small set of songs/creators (HHI, top-N share, Gini coefficient).\n\n**Path parameter:** `run_id`.\n**Auth:** Bearer JWT — caller must be a member of the run's org.\n**Response:** `{ hhi, top10_share_pct, gini, top_concentrations: [...] }`.",
)
def get_underwriting_concentration(
    run_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    run = db.query(UnderwritingRun).filter(
        UnderwritingRun.id == run_id,
        UnderwritingRun.organization_id == membership.organization_id,
    ).first()
    if not run:
        raise HTTPException(status_code=404, detail="Underwriting run not found")

    return run.concentration_data or {}


@router.get(
    "/underwriting/latest",
    summary='Get the latest underwriting run for the org',
    description="Convenience endpoint that returns the most recent completed run headline so dashboards don't have to list-then-fetch.\n\n**Query:** `org_id`.\n**Auth:** Bearer JWT — caller must be a member of the org.\n**Response:** same shape as `/runs/{run_id}` or `null` if no run exists yet.",
)
def get_latest_underwriting(
    scope_creator_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    if scope_creator_id is not None:
        from ..models.models import Creator as _Creator
        belongs = db.query(_Creator.id).filter(
            _Creator.id == scope_creator_id,
            _Creator.organization_id == membership.organization_id,
        ).first()
        if not belongs:
            raise HTTPException(status_code=404, detail="Creator not found in this org")

    q = db.query(UnderwritingRun).filter(
        UnderwritingRun.organization_id == membership.organization_id,
        UnderwritingRun.status == "COMPLETED",
    )
    if scope_creator_id is not None:
        q = q.filter(UnderwritingRun.scope_creator_id == scope_creator_id)
    else:
        q = q.filter(UnderwritingRun.scope_creator_id.is_(None))
    run = q.order_by(UnderwritingRun.created_at.desc()).first()

    if not run:
        return {"has_data": False}

    return {
        "has_data": True,
        "run_id": run.id,
        "status": run.status,
        "kb_version": run.kb_version,
        "inputs": run.inputs,
        "portfolio_summary": run.outputs,
        "spine": run.spine_data,
        "decay": run.decay_data,
        "concentration": run.concentration_data,
        "projections": run.projection_data,
        "valuation": run.valuation_data,
        "exceptions": run.exceptions,
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "completed_at": run.completed_at.isoformat() if run.completed_at else None,
    }


@router.get(
    "/statements/{statement_id}/reconciliation",
    summary='Get the reconciliation snapshot tied to a statement',
    description="Returns the underwriting-side reconciliation for a specific RoyaltyStatement (matched vs. unmatched cashflow, valuation adjustments).\n\n**Path parameter:** `statement_id`.\n**Auth:** Bearer JWT — caller must be a member of the statement's org.\n**Response:** `{ statement_id, reported_total, matched_total, valuation_delta }`.",
)
def get_statement_reconciliation(
    statement_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    result = run_reconciliation_controls(db, statement_id, membership.organization_id)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    db.commit()
    return result


# ---------------------------------------------------------------------------
# Source-typed valuation engine (Task #162)
# ---------------------------------------------------------------------------

class SourceTypedRunRequest(BaseModel):
    scope_creator_id: Optional[int] = None
    scope_song_ids: Optional[List[int]] = None


def _scope_check_creator(db: Session, creator_id: int, org_id: int) -> None:
    from ..models.models import Creator as _Creator
    belongs = db.query(_Creator.id).filter(
        _Creator.id == creator_id,
        _Creator.organization_id == org_id,
    ).first()
    if not belongs:
        raise HTTPException(status_code=404, detail="Creator not found in this org")


def _serialize_source_typed_summary(
    org_id: int,
    db: Session,
    scope_creator_id: Optional[int],
    *,
    summary: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """If ``summary`` is supplied (fresh run) return its aggregated
    shape directly. Otherwise re-aggregate the most recent
    source-typed ``ValuationCalculation`` row per song so the
    ``/summary`` endpoint can serve the latest persisted result
    without recomputing.
    """
    if summary is not None:
        return {
            "computed_at": summary["computed_at"],
            "song_count": summary["song_count"],
            "songs_with_revenue": summary["songs_with_revenue"],
            "by_bucket": summary["by_bucket"],
            "total_annual_revenue_cents": summary["total_annual_revenue_cents"],
            "total_value_cents": summary["total_value_cents"],
            "artist_total_value_cents": summary["artist_total_value_cents"],
            "publisher_total_value_cents": summary["publisher_total_value_cents"],
            "scope": summary["scope"],
            "fresh": True,
        }

    # Re-aggregate from the latest source-typed row per song.
    q = db.query(ValuationCalculation).filter(
        ValuationCalculation.organization_id == org_id,
        ValuationCalculation.valuation_method == "SOURCE_TYPED",
    )

    scoped_song_ids: Optional[set] = None
    if scope_creator_id is not None:
        from ..models.models import SongCredit as _SongCredit
        scoped_song_ids = {
            sid for (sid,) in db.query(_SongCredit.song_id)
            .filter(_SongCredit.creator_id == scope_creator_id)
            .distinct()
            .all()
        }
        if scoped_song_ids:
            q = q.filter(ValuationCalculation.song_id.in_(scoped_song_ids))
        else:
            q = q.filter(False)

    rows = q.order_by(ValuationCalculation.calculation_date.desc()).all()
    latest_per_song: Dict[int, ValuationCalculation] = {}
    for r in rows:
        if r.song_id not in latest_per_song:
            latest_per_song[r.song_id] = r

    if not latest_per_song:
        return {
            "computed_at": None,
            "song_count": 0,
            "songs_with_revenue": 0,
            "by_bucket": {
                b: {"revenue_cents": 0, "multiplier": None, "value_cents": 0}
                for b in ("performance", "mechanical", "sync", "streaming", "other")
            },
            "total_annual_revenue_cents": 0,
            "total_value_cents": 0,
            "artist_total_value_cents": 0,
            "publisher_total_value_cents": 0,
            "scope": {"creator_id": scope_creator_id, "song_ids": None},
            "fresh": False,
        }

    by_bucket: Dict[str, Dict[str, Any]] = {
        b: {"revenue_cents": 0, "value_cents": 0, "multiplier": None}
        for b in ("performance", "mechanical", "sync", "streaming", "other")
    }
    total_value = 0
    total_revenue = 0
    artist_total = 0
    publisher_total = 0
    songs_with_revenue = 0
    latest_calc = None

    for r in latest_per_song.values():
        # Multiplier values are constants per bucket; pick from any row.
        if by_bucket["performance"]["multiplier"] is None and r.multiplier_performance is not None:
            by_bucket["performance"]["multiplier"] = r.multiplier_performance
            by_bucket["mechanical"]["multiplier"] = r.multiplier_mechanical
            by_bucket["sync"]["multiplier"] = r.multiplier_sync
            by_bucket["streaming"]["multiplier"] = r.multiplier_streaming

        for col, bucket in (
            ("revenue_performance_cents", "performance"),
            ("revenue_mechanical_cents", "mechanical"),
            ("revenue_sync_cents", "sync"),
            ("revenue_streaming_cents", "streaming"),
            ("revenue_other_cents", "other"),
        ):
            v = getattr(r, col, 0) or 0
            by_bucket[bucket]["revenue_cents"] += int(v)

        for bucket in ("performance", "mechanical", "sync", "streaming"):
            mult = by_bucket[bucket]["multiplier"] or 0
            rev = getattr(
                r, f"revenue_{bucket}_cents"
            ) or 0
            by_bucket[bucket]["value_cents"] += int(round(rev * mult))

        total_value += int(r.final_valuation_cents or 0)
        total_revenue += int(r.annual_revenue_cents or 0)
        artist_total += int(r.artist_valuation_cents or 0)
        publisher_total += int(r.publisher_valuation_cents or 0)
        if (r.annual_revenue_cents or 0) > 0:
            songs_with_revenue += 1
        if latest_calc is None or (r.calculation_date and r.calculation_date > latest_calc):
            latest_calc = r.calculation_date

    return {
        "computed_at": latest_calc.isoformat() if latest_calc else None,
        "song_count": len(latest_per_song),
        "songs_with_revenue": songs_with_revenue,
        "by_bucket": by_bucket,
        "total_annual_revenue_cents": total_revenue,
        "total_value_cents": total_value,
        "artist_total_value_cents": artist_total,
        "publisher_total_value_cents": publisher_total,
        "scope": {"creator_id": scope_creator_id, "song_ids": None},
        "fresh": False,
    }


@router.post(
    "/source-typed/run",
    summary="Run the source-typed valuation engine for the org (or scoped to a creator/song list)",
    description="Computes annualized revenue per right-category bucket from matched royalty statements, applies industry multipliers (performance 10x, mechanical 9x, sync 7x, streaming 12.5x), and splits each song's value into artist (MASTER) vs publisher (PUBLISHING) shares from RightsSplit rows. Persists one ValuationCalculation row per song with valuation_method='SOURCE_TYPED'.\n\n**Body:** `{ scope_creator_id?, scope_song_ids?: int[] }`.\n**Auth:** Bearer JWT — caller must be a member of the org.",
)
def run_source_typed_valuation(
    request: SourceTypedRunRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    if request.scope_creator_id is not None:
        _scope_check_creator(db, request.scope_creator_id, membership.organization_id)

    try:
        summary = compute_source_typed_valuation(
            db,
            org_id=membership.organization_id,
            scope_creator_id=request.scope_creator_id,
            scope_song_ids=request.scope_song_ids,
            persist=True,
        )
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Source-typed valuation failed: {e}")

    return _serialize_source_typed_summary(
        membership.organization_id,
        db,
        request.scope_creator_id,
        summary=summary,
    )


@router.get(
    "/source-typed/summary",
    summary="Get the latest persisted source-typed valuation breakdown",
    description="Aggregates the most recent source-typed ValuationCalculation row per song into a single org-wide (or per-creator) breakdown. Useful for dashboards that want the current numbers without re-running the engine.\n\n**Query:** `scope_creator_id?`.\n**Auth:** Bearer JWT — caller must be a member of the org.",
)
def get_source_typed_summary(
    scope_creator_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    if scope_creator_id is not None:
        _scope_check_creator(db, scope_creator_id, membership.organization_id)

    return _serialize_source_typed_summary(
        membership.organization_id,
        db,
        scope_creator_id,
    )


# ---------------------------------------------------------------------------
# Phase 5 — Full valuation: Income + Market-Comparable + DCF + Blended
# ---------------------------------------------------------------------------

class FullValuationRunRequest(BaseModel):
    scope_creator_id: Optional[int] = None
    scope_song_ids: Optional[List[int]] = None


def _aggregate_persisted_blended(
    org_id: int,
    db: Session,
    scope_creator_id: Optional[int],
    method: str = "blended",
) -> Dict[str, Any]:
    """Re-aggregate the most recent BLENDED ValuationCalculation row per
    song into the same shape as the in-memory orchestrator returns, so
    GET /full/summary serves stored data without recomputing.

    ``method`` selects the headline number:
        * income            -> revenue_multiple_value_cents
        * market_comparable -> streaming_multiple_value_cents
                              (also stored as market_comp_value_cents)
        * dcf               -> black_box_value_cents
        * blended (default) -> final_valuation_cents
    """
    q = db.query(ValuationCalculation).filter(
        ValuationCalculation.organization_id == org_id,
        ValuationCalculation.valuation_method == "BLENDED",
    )
    if scope_creator_id is not None:
        scoped_song_ids = {
            sid for (sid,) in db.query(SongCredit.song_id)
            .filter(SongCredit.creator_id == scope_creator_id)
            .distinct()
            .all()
        }
        if not scoped_song_ids:
            return _empty_full_summary(org_id, scope_creator_id)
        q = q.filter(ValuationCalculation.song_id.in_(scoped_song_ids))

    rows = q.order_by(ValuationCalculation.calculation_date.desc()).all()
    latest_per_song: Dict[int, ValuationCalculation] = {}
    for r in rows:
        # Scope-pollution guard (mirrors the /full/trend endpoint).
        # BLENDED snapshots are persisted into a single pool keyed by
        # song, so creator-scoped runs (`/full/run?scope_creator_id=X`)
        # would otherwise leak into a later org-wide read because their
        # rows can be the "latest" for a given song. Reject rows whose
        # `calc_metadata.scope_mode` does not match the requested scope:
        #   * org-wide read  -> only accept org-scoped (or untagged) rows
        #   * creator-scoped -> only accept org-scoped rows OR creator
        #                       rows tagged for *this* creator.
        meta = r.calc_metadata or {}
        row_scope_creator = meta.get("scope_creator_id")
        row_scope_mode = meta.get("scope_mode") or (
            "creator" if row_scope_creator else "org"
        )
        if scope_creator_id is None:
            if row_scope_mode != "org":
                continue
        else:
            if row_scope_mode == "creator" and row_scope_creator != scope_creator_id:
                continue
        if r.song_id not in latest_per_song:
            latest_per_song[r.song_id] = r

    if not latest_per_song:
        return _empty_full_summary(org_id, scope_creator_id)

    income_total = market_total = dcf_total = blended_total = 0
    annual_revenue_total = 0
    by_bucket_cents: Dict[str, int] = {b: 0 for b in ("performance", "mechanical", "sync", "streaming", "other")}
    by_bucket_value_cents: Dict[str, int] = {b: 0 for b in ("performance", "mechanical", "sync", "streaming", "other")}
    artist_total = publisher_total = 0
    songs_with_statements = 0
    songs_with_streaming = 0
    confidence_sum = 0.0
    latest_calc = None
    songs_meta_ids = list(latest_per_song.keys())

    for r in latest_per_song.values():
        income_total += int(r.revenue_multiple_value_cents or 0)
        market_total += int(r.streaming_multiple_value_cents or 0)
        dcf_total += int(r.black_box_value_cents or 0)
        blended_total += int(r.final_valuation_cents or 0)
        annual_revenue_total += int(r.annual_revenue_cents or 0)
        artist_total += int(r.artist_valuation_cents or 0)
        publisher_total += int(r.publisher_valuation_cents or 0)
        for col, b in (
            ("revenue_performance_cents", "performance"),
            ("revenue_mechanical_cents", "mechanical"),
            ("revenue_sync_cents", "sync"),
            ("revenue_streaming_cents", "streaming"),
            ("revenue_other_cents", "other"),
        ):
            v = getattr(r, col, 0) or 0
            by_bucket_cents[b] += int(v)
        meta = r.calc_metadata or {}
        ds = meta.get("data_sources") or []
        if "matched_royalty_statements" in ds:
            songs_with_statements += 1
        if "song_streaming_metrics" in ds:
            songs_with_streaming += 1
        conf = meta.get("confidence") or {}
        confidence_sum += float(conf.get("score") or 0.0)
        if latest_calc is None or (r.calculation_date and r.calculation_date > latest_calc):
            latest_calc = r.calculation_date

    # Recompute per-bucket value at the same multipliers we stored on
    # the row so the source-typed view stays internally consistent.
    from ..services.valuation_engine import _BUCKET_MULTIPLIERS
    for b, mult in _BUCKET_MULTIPLIERS.items():
        by_bucket_value_cents[b] = int(round(by_bucket_cents[b] * mult))

    song_count = len(latest_per_song)
    avg_conf = confidence_sum / song_count if song_count else 0.0
    if avg_conf >= 0.66:
        conf_label = "high"
    elif avg_conf >= 0.33:
        conf_label = "medium"
    else:
        conf_label = "low"

    # Top songs by selected method
    method_to_col = {
        "income": "revenue_multiple_value_cents",
        "market_comparable": "streaming_multiple_value_cents",
        "dcf": "black_box_value_cents",
        "blended": "final_valuation_cents",
    }
    metric_col = method_to_col.get(method, "final_valuation_cents")
    meta_rows = (
        db.query(Song.id, Song.title, Song.primary_artist)
        .filter(Song.id.in_(songs_meta_ids))
        .all()
    )
    meta_by_id = {sid: (title, artist) for sid, title, artist in meta_rows}
    songs_summary: List[Dict[str, Any]] = []
    for sid, r in latest_per_song.items():
        title, primary_artist = meta_by_id.get(sid, (None, None))
        songs_summary.append({
            "song_id": sid,
            "title": title,
            "primary_artist": primary_artist,
            "income_base": int(r.revenue_multiple_value_cents or 0) / 100.0,
            "market_base": int(r.streaming_multiple_value_cents or 0) / 100.0,
            "dcf_base": int(r.black_box_value_cents or 0) / 100.0,
            "blended_base": int(r.final_valuation_cents or 0) / 100.0,
            "annual_revenue": int(r.annual_revenue_cents or 0) / 100.0,
            "confidence_score": (r.calc_metadata or {}).get("confidence", {}).get("score") or 0.0,
            "confidence_label": (r.calc_metadata or {}).get("confidence", {}).get("label") or "low",
            "data_sources": (r.calc_metadata or {}).get("data_sources") or [],
            "_metric_cents": int(getattr(r, metric_col, 0) or 0),
        })
    songs_summary.sort(key=lambda s: s["_metric_cents"], reverse=True)
    top_songs = [{k: v for k, v in s.items() if k != "_metric_cents"} for s in songs_summary[:10]]

    # Per-creator share (org-wide only). Uses RightsSplit-weighted
    # attribution via the engine helper so persisted summaries match the
    # economically correct allocation a fresh run produces.
    per_creator_share: List[Dict[str, Any]] = []
    blended_lookup = {s["song_id"]: s["blended_base"] for s in songs_summary}
    if scope_creator_id is None and songs_meta_ids:
        from ..services.valuation_engine import _attribute_songs_to_creators
        attribution = _attribute_songs_to_creators(db, songs_meta_ids)
        creator_totals: Dict[int, Dict[str, Any]] = {}
        for sid, allocations in attribution.items():
            song_value = blended_lookup.get(sid, 0.0)
            if not allocations or song_value <= 0:
                continue
            for cid, cname, share_fraction in allocations:
                if cid not in creator_totals:
                    creator_totals[cid] = {
                        "creator_id": cid,
                        "creator_name": cname or f"Creator #{cid}",
                        "blended_value": 0.0,
                        "song_count": 0,
                    }
                creator_totals[cid]["blended_value"] += song_value * share_fraction
                creator_totals[cid]["song_count"] += 1
        per_creator_share = sorted(
            creator_totals.values(), key=lambda r: r["blended_value"], reverse=True
        )[:25]
        # Total used for share %: sum of blended for the persisted snapshot rows.
        catalog_blended_total = sum(blended_lookup.values()) if blended_lookup else 0.0
        for r in per_creator_share:
            r["blended_value"] = round(r["blended_value"], 2)
            r["blended_base"] = r["blended_value"]
            r["share_pct"] = (
                round(r["blended_value"] / catalog_blended_total * 100.0, 2)
                if catalog_blended_total > 0
                else 0.0
            )

    by_source = {
        b: {
            "revenue_cents": int(by_bucket_cents.get(b, 0)),
            "value_cents": int(by_bucket_value_cents.get(b, 0)),
            "multiplier": _BUCKET_MULTIPLIERS.get(b),
        }
        for b in ("performance", "mechanical", "sync", "streaming", "other")
    }

    from ..services.valuation_engine import _BLEND_WEIGHTS
    return {
        "org_id": org_id,
        "scope": {"creator_id": scope_creator_id, "song_ids": None},
        "computed_at": latest_calc.isoformat() if latest_calc else None,
        "song_count": song_count,
        "songs_with_statements": songs_with_statements,
        "songs_with_streaming": songs_with_streaming,
        "by_methodology": {
            "income": {"low": round(income_total / 100 * 0.85, 2), "base": round(income_total / 100, 2), "high": round(income_total / 100 * 1.15, 2)},
            "market_comparable": {"low": round(market_total / 100 * 0.85, 2), "base": round(market_total / 100, 2), "high": round(market_total / 100 * 1.15, 2)},
            "dcf": {"low": round(dcf_total / 100 * 0.80, 2), "base": round(dcf_total / 100, 2), "high": round(dcf_total / 100 * 1.20, 2)},
            "blended": {"low": round(blended_total / 100 * 0.85, 2), "base": round(blended_total / 100, 2), "high": round(blended_total / 100 * 1.15, 2)},
        },
        "by_source": by_source,
        "annual_revenue_cents": annual_revenue_total,
        "artist_total_value_cents": artist_total,
        "publisher_total_value_cents": publisher_total,
        "weights": dict(_BLEND_WEIGHTS),
        "data_quality": {
            "song_count": song_count,
            "songs_with_statements": songs_with_statements,
            "songs_with_streaming": songs_with_streaming,
            "pct_with_statements": round(songs_with_statements / song_count * 100, 1) if song_count else 0.0,
            "pct_with_streaming": round(songs_with_streaming / song_count * 100, 1) if song_count else 0.0,
            "average_confidence": round(avg_conf, 4),
            "confidence_label": conf_label,
        },
        "top_songs": top_songs,
        "per_creator_share": per_creator_share,
        "selected_method": method,
        "fresh": False,
    }


def _empty_full_summary(org_id: int, scope_creator_id: Optional[int]) -> Dict[str, Any]:
    from ..services.valuation_engine import _BLEND_WEIGHTS, _BUCKET_MULTIPLIERS
    return {
        "org_id": org_id,
        "scope": {"creator_id": scope_creator_id, "song_ids": None},
        "computed_at": None,
        "song_count": 0,
        "songs_with_statements": 0,
        "songs_with_streaming": 0,
        "by_methodology": {
            k: {"low": 0.0, "base": 0.0, "high": 0.0}
            for k in ("income", "market_comparable", "dcf", "blended")
        },
        "by_source": {
            b: {"revenue_cents": 0, "value_cents": 0, "multiplier": _BUCKET_MULTIPLIERS.get(b)}
            for b in ("performance", "mechanical", "sync", "streaming", "other")
        },
        "annual_revenue_cents": 0,
        "artist_total_value_cents": 0,
        "publisher_total_value_cents": 0,
        "weights": dict(_BLEND_WEIGHTS),
        "data_quality": {
            "song_count": 0,
            "songs_with_statements": 0,
            "songs_with_streaming": 0,
            "pct_with_statements": 0.0,
            "pct_with_streaming": 0.0,
            "average_confidence": 0.0,
            "confidence_label": "low",
        },
        "top_songs": [],
        "per_creator_share": [],
        "selected_method": "blended",
        "fresh": False,
    }


from ..services.valuation_engine import _BUCKET_MULTIPLIERS  # used by _aggregate_persisted_blended


@router.post(
    "/full/run",
    summary="Run the blended (Income + Market + DCF) valuation engine",
    description="Computes the per-song full valuation (Income source-typed + Market-Comparable streams × tier band × 10x + DCF over historical statements) and blends them 40/30/30. Persists one BLENDED ValuationCalculation row per scoped song. Returns aggregated by-methodology / by-source totals + top songs + per-creator share + data-quality summary.\n\n**Body:** `{ scope_creator_id?, scope_song_ids?: int[] }`. **Auth:** Bearer JWT.",
)
def run_full_valuation(
    request: FullValuationRunRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")

    if request.scope_creator_id is not None:
        _scope_check_creator(db, request.scope_creator_id, membership.organization_id)

    try:
        summary = compute_full_catalog_valuation(
            db,
            org_id=membership.organization_id,
            scope_creator_id=request.scope_creator_id,
            scope_song_ids=request.scope_song_ids,
            persist=True,
        )
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Full valuation failed: {e}")
    summary["selected_method"] = "blended"
    return summary


@router.get(
    "/full/summary",
    summary="Get the most recent blended valuation breakdown for the org",
    description="Re-aggregates the latest BLENDED ValuationCalculation row per song. Cheap (no recompute). Use `?method=` to switch the headline numbers between `income`, `market_comparable`, `dcf`, and `blended` (default).",
)
def get_full_valuation_summary(
    scope_creator_id: Optional[int] = None,
    method: str = "blended",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")
    if scope_creator_id is not None:
        _scope_check_creator(db, scope_creator_id, membership.organization_id)

    valid_methods = ("income", "market_comparable", "dcf", "blended")
    if method not in valid_methods:
        raise HTTPException(
            status_code=400,
            detail=f"method must be one of {valid_methods}",
        )

    return _aggregate_persisted_blended(
        membership.organization_id, db, scope_creator_id, method=method
    )


@router.get(
    "/full/trend",
    summary="Get historical blended valuation trend",
    description="Returns the catalog's blended valuation per calculation_date (monthly buckets), going back ``months`` months. Used by the historical trend line on the Valuation page.",
)
def get_full_valuation_trend(
    scope_creator_id: Optional[int] = None,
    months: int = 12,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")
    if scope_creator_id is not None:
        _scope_check_creator(db, scope_creator_id, membership.organization_id)

    months = max(1, min(months, 36))
    cutoff = datetime.utcnow() - timedelta(days=months * 31)

    q = db.query(ValuationCalculation).filter(
        ValuationCalculation.organization_id == membership.organization_id,
        ValuationCalculation.valuation_method == "BLENDED",
        ValuationCalculation.calculation_date >= cutoff,
    )
    if scope_creator_id is not None:
        scoped_song_ids = {
            sid for (sid,) in db.query(SongCredit.song_id)
            .filter(SongCredit.creator_id == scope_creator_id)
            .distinct()
            .all()
        }
        if not scoped_song_ids:
            return {"trend": [], "scope": {"creator_id": scope_creator_id}, "months": months}
        q = q.filter(ValuationCalculation.song_id.in_(scoped_song_ids))

    # Trend is a sequence of *snapshots*, not a cumulative sum across runs.
    # Each call to ``/full/run`` writes one row per song with a shared
    # ``calculation_date`` timestamp — that exact timestamp identifies the
    # snapshot. For each day-bucket we therefore:
    #   1. Group rows by their exact ``calculation_date`` (the run id).
    #   2. Sum per-leg cents within each snapshot.
    #   3. Keep only the LATEST snapshot's totals per day, so re-running on
    #      the same day overwrites instead of double-counting.
    # Snapshots also carry a scope tag in ``calc_metadata.scope_mode`` /
    # ``scope_creator_id`` — we discard rows that don't match the
    # requested scope so a creator-scoped subset run cannot collapse an
    # org-wide trend bucket to a partial total. Rows from older runs
    # (pre-scope-tag) are treated as ``org`` for backwards compatibility.
    # We keep this in Python so SQLite unit tests (which lack date_trunc)
    # still pass.
    rows = q.order_by(ValuationCalculation.calculation_date.asc()).all()
    by_snapshot: Dict[Any, Dict[str, int]] = {}
    snapshot_day: Dict[Any, str] = {}
    for r in rows:
        if not r.calculation_date:
            continue
        # Scope-match filter
        meta = r.calc_metadata or {}
        row_scope_creator = meta.get("scope_creator_id")
        row_scope_mode = meta.get("scope_mode") or ("creator" if row_scope_creator else "org")
        if scope_creator_id is None:
            # Org-wide trend: only org-wide snapshots count.
            if row_scope_mode != "org":
                continue
        else:
            # Creator-scoped trend: only this creator's snapshots count.
            # (The query already restricted song_ids; this prevents an
            #  org-wide run from being mistaken for a scoped one.)
            if row_scope_mode != "creator" or row_scope_creator != scope_creator_id:
                continue
        snap_key = r.calculation_date  # exact run timestamp
        if snap_key not in by_snapshot:
            by_snapshot[snap_key] = {
                "blended_cents": 0,
                "income_cents": 0,
                "market_cents": 0,
                "dcf_cents": 0,
            }
            snapshot_day[snap_key] = r.calculation_date.date().isoformat()
        by_snapshot[snap_key]["blended_cents"] += int(r.final_valuation_cents or 0)
        by_snapshot[snap_key]["income_cents"] += int(r.revenue_multiple_value_cents or 0)
        by_snapshot[snap_key]["market_cents"] += int(r.streaming_multiple_value_cents or 0)
        by_snapshot[snap_key]["dcf_cents"] += int(r.black_box_value_cents or 0)

    # Pick the latest snapshot per day (snapshots are timestamped, so
    # within a given day the largest timestamp wins).
    by_day: Dict[str, Dict[str, int]] = {}
    for snap_key in sorted(by_snapshot.keys()):
        day = snapshot_day[snap_key]
        by_day[day] = by_snapshot[snap_key]  # latest snapshot wins

    trend = [
        {
            "date": d,
            "blended": by_day[d]["blended_cents"] / 100.0,
            "income": by_day[d]["income_cents"] / 100.0,
            "market_comparable": by_day[d]["market_cents"] / 100.0,
            "dcf": by_day[d]["dcf_cents"] / 100.0,
        }
        for d in sorted(by_day.keys())
    ]
    return {
        "scope": {"creator_id": scope_creator_id},
        "months": months,
        "trend": trend,
    }


@router.get(
    "/report/pdf",
    summary="Download a styled blended valuation report as a PDF",
    description="ReportLab-generated multi-page PDF: cover, executive summary (blended NPV band + confidence), methodology breakdown (Income / Market / DCF), revenue-by-source table, top-10 songs, data-sources & disclaimer. Honors `?scope_creator_id=` for per-client decks.",
)
def download_full_valuation_pdf(
    scope_creator_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id
    ).first()
    if not membership:
        raise HTTPException(status_code=404, detail="Organization membership not found")
    if scope_creator_id is not None:
        _scope_check_creator(db, scope_creator_id, membership.organization_id)

    org = db.query(Organization).filter(
        Organization.id == membership.organization_id
    ).first()
    org_name = org.name if org else "Your Organization"
    creator_name = None
    if scope_creator_id is not None:
        c = db.query(Creator).filter(Creator.id == scope_creator_id).first()
        creator_name = c.display_name if c else f"Creator #{scope_creator_id}"

    summary = _aggregate_persisted_blended(
        membership.organization_id, db, scope_creator_id, method="blended"
    )

    pdf_bytes = _build_valuation_pdf(summary, org_name, creator_name)
    filename = (
        f"cadence_valuation_{(creator_name or 'catalog').lower().replace(' ', '_')}_"
        f"{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_valuation_pdf(
    summary: Dict[str, Any],
    org_name: str,
    creator_name: Optional[str] = None,
) -> bytes:
    """Render a multi-page blended-valuation PDF in the Cadence sage palette.

    Sections (in order):
      1. Cover page — title, scope, blended NPV band.
      2. Executive summary — by-methodology table with bands + confidence pill.
      3. Revenue-by-source — bucket table with multipliers.
      4. Top-10 songs by blended value.
      5. Data sources & disclaimer.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )

    sage = colors.HexColor("#5B8A72")
    sage_light = colors.HexColor("#E5EEDF")
    dark_text = colors.HexColor("#3D4A44")
    muted = colors.HexColor("#7A8580")
    border = colors.HexColor("#E0E5E2")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.75 * inch, bottomMargin=0.75 * inch,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
        title=f"Cadence Valuation — {org_name}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontSize=24,
        textColor=dark_text, leading=28, spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"], fontSize=12,
        textColor=muted, leading=15, spaceAfter=18,
    )
    h2 = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontSize=14,
        textColor=sage, leading=18, spaceBefore=14, spaceAfter=8,
    )
    body = ParagraphStyle(
        "Body", parent=styles["Normal"], fontSize=10,
        textColor=dark_text, leading=14, spaceAfter=6,
    )
    small = ParagraphStyle(
        "Small", parent=styles["Normal"], fontSize=8,
        textColor=muted, leading=11, spaceAfter=4,
    )

    elements = []

    # ---- Cover ------------------------------------------------------------
    elements.append(Paragraph("Catalog Valuation Report", title_style))
    scope_str = f"{org_name}"
    if creator_name:
        scope_str += f"  ·  Scoped to {creator_name}"
    elements.append(Paragraph(scope_str, subtitle_style))
    elements.append(Paragraph(
        f"Generated {datetime.utcnow().strftime('%B %d, %Y')} by Cadence Catalog Intelligence",
        small,
    ))
    elements.append(Spacer(1, 0.3 * inch))

    bm = summary.get("by_methodology", {})
    blended = bm.get("blended", {"low": 0, "base": 0, "high": 0})
    fmt = lambda v: f"${(v or 0):,.0f}"
    hero_data = [
        [Paragraph("<b>Blended Catalog Value (NPV)</b>", body)],
        [Paragraph(f"<font size=22 color='#5B8A72'><b>{fmt(blended.get('base'))}</b></font>", body)],
        [Paragraph(
            f"<font color='#7A8580' size=9>Range: {fmt(blended.get('low'))} – {fmt(blended.get('high'))}</font>",
            body,
        )],
    ]
    hero = Table(hero_data, colWidths=[6.5 * inch])
    hero.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), sage_light),
        ("BOX", (0, 0), (-1, -1), 0.5, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 18),
        ("RIGHTPADDING", (0, 0), (-1, -1), 18),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    elements.append(hero)

    dq = summary.get("data_quality", {})
    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph(
        f"<b>Confidence:</b> {dq.get('confidence_label', 'low').upper()}  ·  "
        f"<b>Songs:</b> {dq.get('song_count', 0)}  ·  "
        f"<b>With statements:</b> {dq.get('pct_with_statements', 0)}%  ·  "
        f"<b>With streaming data:</b> {dq.get('pct_with_streaming', 0)}%",
        body,
    ))

    # ---- Section 1: Executive Methodology Breakdown -----------------------
    elements.append(Paragraph("Methodology Breakdown", h2))
    elements.append(Paragraph(
        "Each methodology is computed independently from the same underlying data, then "
        "blended at <b>40% Income / 30% Market-Comparable / 30% DCF</b>. "
        "Bands reflect low / base / high scenarios.",
        body,
    ))
    method_rows = [["Methodology", "Low", "Base", "High", "Weight"]]
    weights = summary.get("weights", {"income": 0.4, "market_comparable": 0.3, "dcf": 0.3})
    for key, label in (
        ("income", "Income (source-typed)"),
        ("market_comparable", "Market Comparable"),
        ("dcf", "Discounted Cash Flow"),
        ("blended", "Blended"),
    ):
        seg = bm.get(key, {})
        wt = weights.get(key, "—") if key != "blended" else "100%"
        if isinstance(wt, float):
            wt = f"{int(wt * 100)}%"
        method_rows.append([
            label,
            fmt(seg.get("low")),
            fmt(seg.get("base")),
            fmt(seg.get("high")),
            str(wt),
        ])
    mt = Table(method_rows, colWidths=[2.4*inch, 1.0*inch, 1.0*inch, 1.0*inch, 0.7*inch])
    mt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), sage),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TEXTCOLOR", (0, 1), (-1, -1), dark_text),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, sage_light]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, -2), (-1, -2), 0.5, border),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(mt)

    # ---- Section 2: Revenue by Source -------------------------------------
    elements.append(Paragraph("Revenue by Source", h2))
    elements.append(Paragraph(
        "Annualized revenue from matched royalty statements, bucketed by canonical right "
        "category and valued at industry multiples.",
        body,
    ))
    src_rows = [["Source", "Annual Revenue", "Multiplier", "Contribution to Value"]]
    bs = summary.get("by_source", {})
    for key, label in (
        ("performance", "Performance (PROs)"),
        ("mechanical", "Mechanical (MLC / HFA)"),
        ("sync", "Sync"),
        ("streaming", "Streaming (DSPs)"),
        ("other", "Other / Unclassified"),
    ):
        seg = bs.get(key, {})
        rev = (seg.get("revenue_cents") or 0) / 100.0
        val = (seg.get("value_cents") or 0) / 100.0
        mult = seg.get("multiplier")
        src_rows.append([
            label,
            fmt(rev),
            f"{mult}×" if mult else "—",
            fmt(val),
        ])
    st = Table(src_rows, colWidths=[2.4*inch, 1.4*inch, 1.0*inch, 1.7*inch])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), sage),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, sage_light]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(st)

    # ---- Section 3: Top Songs ---------------------------------------------
    top_songs = summary.get("top_songs", [])[:10]
    if top_songs:
        elements.append(PageBreak())
        elements.append(Paragraph("Top Songs by Blended Value", h2))
        ts_rows = [["#", "Title", "Artist", "Income", "Market", "DCF", "Blended"]]
        for i, s in enumerate(top_songs, start=1):
            title = (s.get("title") or "—")[:35]
            artist = (s.get("primary_artist") or "—")[:25]
            ts_rows.append([
                str(i),
                title,
                artist,
                fmt(s.get("income_base")),
                fmt(s.get("market_base")),
                fmt(s.get("dcf_base")),
                fmt(s.get("blended_base")),
            ])
        tt = Table(ts_rows, colWidths=[0.3*inch, 2.0*inch, 1.4*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
        tt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), sage),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, sage_light]),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(tt)

    # ---- Section 4: Per-creator share (org-wide only) ---------------------
    pcs = summary.get("per_creator_share", [])[:10]
    if pcs:
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph("Top Contributors by Blended Value", h2))
        pc_rows = [["#", "Creator", "Songs", "Blended Value"]]
        for i, c in enumerate(pcs, start=1):
            pc_rows.append([
                str(i),
                (c.get("creator_name") or "—")[:40],
                str(c.get("song_count", 0)),
                fmt(c.get("blended_value")),
            ])
        pct = Table(pc_rows, colWidths=[0.3*inch, 3.6*inch, 0.8*inch, 1.8*inch])
        pct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), sage),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, sage_light]),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(pct)

    # ---- Section 5: Data Sources -----------------------------------------
    # Aggregate the per-song `data_sources` arrays into a single coverage
    # roll-up, plus the org-level data-quality counts already on `summary`.
    elements.append(Spacer(1, 0.25 * inch))
    elements.append(Paragraph("Data Sources", h2))
    elements.append(Paragraph(
        "Inputs feeding the valuation engines for this run, with coverage "
        "across the catalog. Higher coverage materially improves confidence.",
        body,
    ))
    src_label_map = {
        "matched_royalty_statements": "Matched Royalty Statements (Income / DCF)",
        "song_streaming_metrics": "Song Streaming Metrics (Market-Comparable)",
        "rights_splits": "Rights Splits (Per-Creator Attribution)",
        "song_credits": "Song Credits (Equal-Split Fallback)",
    }
    src_counts: Dict[str, int] = {}
    for s in summary.get("top_songs", []):
        for ds_key in (s.get("data_sources") or []):
            src_counts[ds_key] = src_counts.get(ds_key, 0) + 1
    # Always show the two canonical coverage numbers from data_quality so
    # the table is meaningful even when top_songs is empty.
    ds_rows = [["Data Source", "Songs Covered", "Catalog Coverage"]]
    n_total = max(1, dq.get("song_count", 0) or 0)
    ds_rows.append([
        src_label_map["matched_royalty_statements"],
        str(dq.get("songs_with_statements", 0)),
        f"{dq.get('pct_with_statements', 0)}%",
    ])
    ds_rows.append([
        src_label_map["song_streaming_metrics"],
        str(dq.get("songs_with_streaming", 0)),
        f"{dq.get('pct_with_streaming', 0)}%",
    ])
    # Surface any additional sources we observed in the per-song metadata
    # (rights_splits / song_credits / etc.) that aren't on data_quality.
    for key, label in src_label_map.items():
        if key in ("matched_royalty_statements", "song_streaming_metrics"):
            continue
        n = src_counts.get(key, 0)
        if n:
            ds_rows.append([label, str(n), f"{round(n / n_total * 100, 1)}%"])
    dst = Table(ds_rows, colWidths=[3.4*inch, 1.2*inch, 1.5*inch])
    dst.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), sage),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, sage_light]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(dst)

    # ---- Disclaimer -------------------------------------------------------
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph("Methodology & Disclaimer", h2))
    elements.append(Paragraph(
        "<b>Income</b> annualizes matched royalty-statement revenue by source type and "
        "applies industry multiples (performance 10×, mechanical 9×, sync 7×, streaming 12.5×). "
        "<b>Market-Comparable</b> annualizes streams from connected DSPs by song age, applies a "
        "tier-band per-stream rate, and multiplies by 10× to mirror recent catalog transactions. "
        "<b>DCF</b> projects historical annual revenue forward using the song's fitted growth/decay "
        "rate, discounts at 10%, and adds a Gordon-growth terminal value (g=2%).",
        small,
    ))
    elements.append(Paragraph(
        "These valuations are model-derived estimates and should not be construed as a binding "
        "bid, appraisal, or financial advice. Final transaction value depends on counterparty "
        "due-diligence, contract structure, and market conditions.",
        small,
    ))

    doc.build(elements)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Spec'd Phase 5 contract: GET /api/organizations/{org_id}/valuation/catalog
# (mirrored at /api/v1/organizations/{org_id}/valuation/catalog by
# main._mount_v1_routes). Single endpoint with method discriminator that
# returns the full per-method catalog summary used by the Valuation page.
# ---------------------------------------------------------------------------

_VALID_METHODS = {"income", "market_comparable", "dcf", "blended"}


@org_router.get(
    "/{org_id}/valuation/catalog",
    summary="Get the full catalog valuation summary for the org",
    description=(
        "Spec'd Phase 5 contract. Returns the same shape as "
        "`/api/valuation/full/summary` (by_methodology, by_source, top_songs, "
        "per_creator_share, data_quality, weights, confidence) plus a "
        "`selected_method` discriminator chosen by the `?method=` query.\n\n"
        "**Path:** `org_id` — must match a current org membership of the caller.\n"
        "**Query:** `creator_id?` (per-creator scope), `method?` "
        "(income | market_comparable | dcf | blended; default = blended), "
        "`refresh?` (true to force a fresh recompute even if a snapshot "
        "already exists; defaults to snapshot-reuse for low latency)."
    ),
)
def get_org_catalog_valuation(
    org_id: int,
    creator_id: Optional[int] = None,
    method: str = "blended",
    refresh: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    method = (method or "blended").lower()
    if method not in _VALID_METHODS:
        raise HTTPException(
            status_code=400,
            detail=f"method must be one of {sorted(_VALID_METHODS)}",
        )

    # Auth: caller must be a member of the requested org.
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id,
        OrganizationMember.organization_id == org_id,
    ).first()
    if not membership:
        raise HTTPException(status_code=403, detail="Not a member of this organization")
    if creator_id is not None:
        _scope_check_creator(db, creator_id, org_id)

    # Default semantics: serve the latest persisted BLENDED snapshot when
    # one exists (low-latency reads). When no snapshot exists OR the caller
    # explicitly passes ?refresh=true, run the full per-method engine over
    # the *entire org catalog* and persist a fresh snapshot before
    # aggregating. The recompute is intentionally org-wide regardless of
    # the optional ?creator_id= filter — the snapshot pool is shared and
    # `_aggregate_persisted_blended()` does the per-creator scoping at
    # read time. Persisting a creator-scoped snapshot here would pollute
    # subsequent org-wide aggregations because every BLENDED row lands in
    # the same `valuation_calculations` pool keyed by song.
    #
    # ``has_snapshot`` must be scope-aware: a pool that contains *only*
    # creator-scoped BLENDED rows (e.g. from prior `/full/run?scope_creator_id`
    # invocations) would not satisfy an org-wide catalog read because the
    # aggregator filters those out by scope tag. Walk the row metadata
    # so we trigger a recompute when the existing pool is unusable for
    # the requested scope.
    has_snapshot = False
    for r in db.query(ValuationCalculation).filter(
        ValuationCalculation.organization_id == org_id,
        ValuationCalculation.valuation_method == "BLENDED",
    ).all():
        meta = r.calc_metadata or {}
        row_scope_creator = meta.get("scope_creator_id")
        row_scope_mode = meta.get("scope_mode") or (
            "creator" if row_scope_creator else "org"
        )
        # Org-wide read needs at least one org-tagged (or untagged) row.
        # Creator-scoped read accepts org-tagged rows (filtered by song
        # at aggregation time) OR a row tagged for *this* creator.
        if creator_id is None:
            if row_scope_mode == "org":
                has_snapshot = True
                break
        else:
            if row_scope_mode == "org" or row_scope_creator == creator_id:
                has_snapshot = True
                break
    if refresh or not has_snapshot:
        try:
            compute_full_catalog_valuation(db, org_id=org_id, persist=True)
            db.commit()
        except Exception:
            db.rollback()
            raise

    summary = _aggregate_persisted_blended(
        org_id=org_id,
        db=db,
        scope_creator_id=creator_id,
        method=method,
    )
    summary["selected_method"] = method
    summary["scope"] = {"org_id": org_id, "creator_id": creator_id}
    return summary


# ---------------------------------------------------------------------------
# Spec'd Phase 5 contract:
#   GET /api/organizations/{org_id}/valuation/report/pdf?creator_id=X
# (mirrored at /api/v1/organizations/{org_id}/valuation/report/pdf).
# Same blended PDF as /api/valuation/report/pdf, with explicit org_id in path
# and `creator_id` query param matching the org-scoped catalog endpoint.
# ---------------------------------------------------------------------------


@org_router.get(
    "/{org_id}/valuation/report/pdf",
    summary="Download the styled blended valuation PDF for an org (or a creator within it)",
    description=(
        "Spec'd Phase 5 contract. ReportLab-generated multi-page PDF: cover, "
        "executive summary (blended NPV band + confidence), methodology "
        "breakdown (Income / Market / DCF), revenue-by-source table, top-10 "
        "songs, data-sources & disclaimer.\n\n"
        "**Path:** `org_id` — must match a current org membership of the caller.\n"
        "**Query:** `creator_id?` — restrict the deck to a single creator in the org."
    ),
)
def download_org_valuation_pdf(
    org_id: int,
    creator_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id,
        OrganizationMember.organization_id == org_id,
    ).first()
    if not membership:
        raise HTTPException(status_code=403, detail="Not a member of this organization")
    if creator_id is not None:
        _scope_check_creator(db, creator_id, org_id)

    org = db.query(Organization).filter(Organization.id == org_id).first()
    org_name = org.name if org else "Your Organization"
    creator_name = None
    if creator_id is not None:
        c = db.query(Creator).filter(Creator.id == creator_id).first()
        creator_name = c.display_name if c else f"Creator #{creator_id}"

    summary = _aggregate_persisted_blended(
        org_id=org_id,
        db=db,
        scope_creator_id=creator_id,
        method="blended",
    )

    pdf_bytes = _build_valuation_pdf(summary, org_name, creator_name)
    filename = (
        f"cadence_valuation_{(creator_name or 'catalog').lower().replace(' ', '_')}_"
        f"{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    )
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
