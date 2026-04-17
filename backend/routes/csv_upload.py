from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import date
import csv
import io
import logging

from ..models import get_db, Song, SongCredit, Creator, OrganizationMember, User, SongChecklistStatus, ChecklistItem
from ..utils.auth import get_current_user
from ..utils.csv_parser import parse_csv_with_ai, apply_mapping_to_rows, validate_mapped_data, infer_mapping_from_data

logger = logging.getLogger("cadence")

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

router = APIRouter(prefix="/api/csv", tags=["CSV Import"])


class ColumnMapping(BaseModel):
    original: str
    mapped_to: Optional[str]


class CSVPreviewResponse(BaseModel):
    headers: List[str]
    mapping: Dict[str, Optional[str]]
    preview_rows: List[Dict[str, str]]
    row_count: int
    success: bool
    error: Optional[str] = None


class MappedSongData(BaseModel):
    title: str
    primary_artist: Optional[str] = None
    isrc: Optional[str] = None
    iswc: Optional[str] = None
    project_title: Optional[str] = None
    release_date: Optional[str] = None
    label: Optional[str] = None
    publishing_percentage: Optional[float] = None
    master_percentage: Optional[float] = None
    advance_amount: Optional[float] = None
    recording_code: Optional[str] = None
    notes: Optional[str] = None


class CSVImportRequest(BaseModel):
    mapping: Dict[str, Optional[str]]
    rows: List[Dict[str, str]]
    creator_id: Optional[int] = None
    create_new_creator: bool = False
    new_creator_name: Optional[str] = None
    staged_file_id: Optional[str] = None
    staged_filename: Optional[str] = None
    staged_mime: Optional[str] = None
    extraction_method: Optional[str] = None
    contract_terms: Optional[Dict[str, Any]] = None
    document_info: Optional[Dict[str, Any]] = None
    is_text_paste: bool = False


class ImportResult(BaseModel):
    songs_created: int
    songs_failed: int
    creator_id: Optional[int]
    errors: List[str]


def format_excel_cell(cell_value, number_format=None) -> str:
    """Format Excel cell value to string, handling dates and percentages specially."""
    if cell_value is None:
        return ""
    
    from datetime import datetime, date as date_type
    
    if isinstance(cell_value, (datetime, date_type)):
        return cell_value.strftime("%Y-%m-%d")
    
    if isinstance(cell_value, (int, float)):
        if number_format and '%' in str(number_format):
            converted = cell_value * 100
            if converted == int(converted):
                return str(int(converted))
            return str(round(converted, 4))
        if cell_value == int(cell_value):
            return str(int(cell_value))
        return str(cell_value)
    
    return str(cell_value).strip()


def is_valid_header_row(row: tuple) -> bool:
    """Check if a row looks like a header row (has meaningful text in multiple cells)."""
    if not row:
        return False
    
    non_empty_count = 0
    text_count = 0
    
    for cell in row:
        if cell is not None and str(cell).strip():
            non_empty_count += 1
            if isinstance(cell, str) and not cell.replace('.', '').replace('-', '').isdigit():
                text_count += 1
    
    return non_empty_count >= 2 and text_count >= 1


def parse_excel_file(content: bytes) -> tuple[list, list]:
    """Parse Excel file and return headers and rows."""
    if not EXCEL_SUPPORT:
        raise HTTPException(status_code=400, detail="Excel support not available. Please upload a CSV file.")
    
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    
    all_rows_values = list(ws.iter_rows(values_only=True))
    all_rows_cells = list(ws.iter_rows(values_only=False))
    
    if not all_rows_values:
        wb.close()
        raise HTTPException(status_code=400, detail="Excel file has no data")
    
    header_row_idx = 0
    for idx, row in enumerate(all_rows_values[:10]):
        if is_valid_header_row(row):
            header_row_idx = idx
            break
    
    header_row = all_rows_values[header_row_idx]
    data_rows_cells = all_rows_cells[header_row_idx + 1:]
    data_rows_values = all_rows_values[header_row_idx + 1:]
    
    headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(header_row)]
    
    generic_header_count = sum(1 for h in headers if h.startswith("Column_"))
    if generic_header_count > len(headers) // 2 and data_rows_values:
        first_data = data_rows_values[0] if data_rows_values else None
        if first_data and is_valid_header_row(first_data):
            headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(first_data)]
            data_rows_cells = data_rows_cells[1:]
    
    rows = []
    for row_cells in data_rows_cells:
        if any(cell.value is not None for cell in row_cells):
            row_dict = {}
            for i, cell in enumerate(row_cells):
                if i < len(headers):
                    row_dict[headers[i]] = format_excel_cell(cell.value, cell.number_format)
            rows.append(row_dict)
    
    wb.close()
    return headers, rows


def parse_csv_file(content: bytes) -> tuple[list, list]:
    """Parse CSV file and return headers and rows."""
    try:
        text_content = content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text_content = content.decode('latin-1')
        except:
            raise HTTPException(status_code=400, detail="Unable to decode file. Please ensure it's a valid CSV.")
    
    reader = csv.DictReader(io.StringIO(text_content))
    headers = reader.fieldnames or []
    
    if not headers:
        raise HTTPException(status_code=400, detail="CSV file has no headers")
    
    rows = list(reader)
    return headers, rows


@router.post("/preview/{org_id}", response_model=CSVPreviewResponse, summary="Preview a CSV import", description="Parses the CSV, runs the AI column mapper, and returns a row-by-row preview with confidence scores. No data is written.")
async def preview_csv(
    org_id: int,
    file: UploadFile = File(...),
    all_rows: bool = Query(False, description="Return all rows instead of just preview"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id,
        OrganizationMember.organization_id == org_id
    ).first()
    
    if not membership:
        raise HTTPException(status_code=403, detail="Not authorized to access this organization")
    
    filename = file.filename.lower() if file.filename else ""
    is_excel = filename.endswith('.xlsx') or filename.endswith('.xls')
    is_csv = filename.endswith('.csv')
    
    if not is_excel and not is_csv:
        raise HTTPException(status_code=400, detail="File must be a CSV or Excel file (.csv, .xlsx, .xls)")
    
    content = await file.read()
    
    if is_excel:
        headers, rows = parse_excel_file(content)
    else:
        headers, rows = parse_csv_file(content)
    
    ai_result = parse_csv_with_ai("", headers, org_id=org_id)
    initial_mapping = ai_result.get("mapping", {})
    
    generic_count = sum(1 for h in headers if h.startswith("Column_"))
    if generic_count > 0 and rows:
        initial_mapping = infer_mapping_from_data(headers, rows, initial_mapping)
    
    preview_rows = rows if all_rows else rows[:5]
    
    return CSVPreviewResponse(
        headers=headers,
        mapping=initial_mapping,
        preview_rows=preview_rows,
        row_count=len(rows),
        success=ai_result.get("success", False),
        error=ai_result.get("error")
    )


_DOC_EXTS = ('.pdf', '.docx', '.doc')
_IMAGE_EXTS = ('.png', '.jpg', '.jpeg', '.webp', '.gif', '.bmp', '.tiff')
_TEXT_EXTS = ('.txt', '.md', '.tsv')
_TABULAR_EXTS = ('.csv', '.xlsx', '.xls')


class TextPreviewRequest(BaseModel):
    text: str


@router.post("/document-preview/{org_id}", summary="Preview a document (PDF/DOCX) import", description="Runs the AI document parser to preview a Schedule A or contract. No data is written.")
async def preview_document(
    org_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from backend.services.document_parser import parse_document_unified

    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id,
        OrganizationMember.organization_id == org_id
    ).first()

    if not membership:
        raise HTTPException(status_code=403, detail="Not authorized to access this organization")

    filename = file.filename or ""
    lower = filename.lower()
    accepted = _DOC_EXTS + _IMAGE_EXTS + _TEXT_EXTS + _TABULAR_EXTS
    if not lower.endswith(accepted):
        raise HTTPException(
            status_code=400,
            detail="File must be a CSV, Excel, PDF, Word, image (PNG/JPG), or plain text document.",
        )

    max_size = 20 * 1024 * 1024 if lower.endswith(_IMAGE_EXTS) else 10 * 1024 * 1024
    if file.size and file.size > max_size:
        raise HTTPException(status_code=400, detail=f"File too large. Maximum size is {max_size // (1024*1024)}MB.")

    content = await file.read()

    # Stash the raw upload so it can be persisted alongside the import for
    # audit / re-extraction. The staged_file_id is opaque and only useful when
    # echoed back to /import.
    from ..services import schedule_a_storage
    staged_id, staged_key = schedule_a_storage.stage_upload(org_id, content, filename)
    sha256 = schedule_a_storage.hash_bytes(content)
    staging_meta = {
        "staged_file_id": staged_id,
        "staged_filename": filename,
        "staged_size": len(content),
        "staged_mime": file.content_type or None,
        "staged_sha256": sha256,
    }

    # Tabular formats (CSV/Excel) go through the existing column-mapping
    # parser, then are wrapped in the unified preview shape so callers see one
    # consistent response regardless of input type.
    if lower.endswith(_TABULAR_EXTS):
        if lower.endswith(('.xlsx', '.xls')):
            headers, rows = parse_excel_file(content)
        else:
            headers, rows = parse_csv_file(content)
        ai_result = parse_csv_with_ai("", headers, org_id=org_id)
        initial_mapping = ai_result.get("mapping", {})
        generic_count = sum(1 for h in headers if h.startswith("Column_"))
        if generic_count > 0 and rows:
            initial_mapping = infer_mapping_from_data(headers, rows, initial_mapping)
        # Tag rows with high deterministic confidence and the spreadsheet source label
        tagged: List[Dict[str, Any]] = []
        for r in rows:
            r2 = dict(r)
            r2.setdefault("_confidence", 0.95)
            r2.setdefault("_source", "spreadsheet row")
            tagged.append(r2)
        resp = {
            "success": True,
            "extraction_method": "tabular_excel" if lower.endswith(('.xlsx', '.xls')) else "tabular_csv",
            "schedule_a_songs": [],
            "schedule_b_songs": [],
            "preview_rows": tagged,
            "preview_total": len(tagged),
            "headers": headers,
            "mapping": initial_mapping,
            "row_count": len(tagged),
            "creator_info": {},
            "contract_terms": {},
            "warnings": [],
            "errors": [],
        }
        resp.update(staging_meta)
        return resp

    result = parse_document_unified(content, filename, org_id=org_id)

    if result.errors and not result.schedule_a_songs and not result.schedule_b_songs:
        # Don't leave the file staged if we're rejecting the upload outright
        try:
            schedule_a_storage.delete(staged_key)
        except Exception:
            pass
        raise HTTPException(status_code=400, detail=result.errors[0])

    payload = result.to_preview_response()
    payload.update(staging_meta)
    return payload


@router.post("/text-preview/{org_id}", summary="Preview pasted-text import", description="Same as document preview but accepts raw pasted text.")
async def preview_pasted_text(
    org_id: int,
    payload: TextPreviewRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from backend.services.document_parser import parse_document_unified

    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id,
        OrganizationMember.organization_id == org_id
    ).first()
    if not membership:
        raise HTTPException(status_code=403, detail="Not authorized to access this organization")

    text = (payload.text or "").strip()
    if len(text) < 5:
        raise HTTPException(status_code=400, detail="Please paste at least a few characters of text.")
    if len(text) > 200_000:
        raise HTTPException(status_code=400, detail="Pasted text is too long (200K char max).")

    from ..services import schedule_a_storage
    encoded = text.encode("utf-8")
    staged_id, staged_key = schedule_a_storage.stage_upload(org_id, encoded, "pasted.txt")
    sha256 = schedule_a_storage.hash_bytes(encoded)
    staging_meta = {
        "staged_file_id": staged_id,
        "staged_filename": "pasted.txt",
        "staged_size": len(encoded),
        "staged_mime": "text/plain",
        "staged_sha256": sha256,
        "is_text_paste": True,
    }

    result = parse_document_unified(None, "pasted.txt", pasted_text=text, org_id=org_id)
    if result.errors and not result.schedule_a_songs and not result.schedule_b_songs:
        try:
            schedule_a_storage.delete(staged_key)
        except Exception:
            pass
        raise HTTPException(status_code=400, detail=result.errors[0])

    out = result.to_preview_response()
    out.update(staging_meta)
    return out


@router.post("/import/{org_id}", response_model=ImportResult, summary="Commit a previewed import", description="Executes the import based on the previously confirmed preview payload. Returns counts of created / updated / skipped rows.")
async def import_csv(
    org_id: int,
    request: CSVImportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    membership = db.query(OrganizationMember).filter(
        OrganizationMember.user_id == current_user.id,
        OrganizationMember.organization_id == org_id
    ).first()
    
    if not membership:
        raise HTTPException(status_code=403, detail="Not authorized to access this organization")
    
    creator = None
    if request.create_new_creator and request.new_creator_name:
        creator = Creator(
            organization_id=org_id,
            display_name=request.new_creator_name,
            roles=["ARTIST"]
        )
        db.add(creator)
        db.flush()
    elif request.creator_id:
        creator = db.query(Creator).filter(
            Creator.id == request.creator_id,
            Creator.organization_id == org_id
        ).first()
        if not creator:
            raise HTTPException(status_code=404, detail="Creator not found")
    
    mapped_rows = apply_mapping_to_rows(request.rows, request.mapping)
    validation = validate_mapped_data(mapped_rows)
    
    songs_created = 0
    errors = []
    created_songs = []
    
    from datetime import datetime
    
    for idx, row_data in enumerate(validation["valid_rows"]):
        try:
            release_date = None
            if row_data.get("release_date"):
                try:
                    release_date = datetime.strptime(row_data["release_date"], "%Y-%m-%d").date()
                except:
                    pass
            
            song = Song(
                organization_id=org_id,
                title=row_data.get("title", "Untitled"),
                primary_artist=row_data.get("primary_artist") or (creator.display_name if creator else "Unknown"),
                isrc=row_data.get("isrc"),
                iswc=row_data.get("iswc"),
                project_title=row_data.get("project_title"),
                release_date=release_date,
                label=row_data.get("label"),
                publishing_percentage=row_data.get("publishing_percentage"),
                master_percentage=row_data.get("master_percentage"),
                advance_amount=row_data.get("advance_amount"),
                recording_code=row_data.get("recording_code"),
                notes=row_data.get("notes"),
                is_released=(release_date is not None),
                status_health_score=0.0
            )
            db.add(song)
            created_songs.append(song)
            songs_created += 1
        except Exception as e:
            errors.append(f"Row {idx + 1}: {str(e)}")
    
    try:
        db.flush()
    except Exception as e:
        db.rollback()
        logger.error(f"CSV import flush failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to save imported songs: {str(e)}")
    
    checklist_items = db.query(ChecklistItem).all()
    checklist_by_code = {item.code: item for item in checklist_items}
    total_weight = sum(item.weight for item in checklist_items) or 1
    
    from ..utils.health_sync import FIELD_TO_CHECKLIST_MAP, NA_CAPABLE_FIELDS
    
    for song in created_songs:
        completed_weight = 0
        for field, code in FIELD_TO_CHECKLIST_MAP.items():
            item = checklist_by_code.get(code)
            if not item:
                continue
            value = getattr(song, field, None)
            if value is None and field in NA_CAPABLE_FIELDS:
                status_val = "NOT_STARTED"
            elif isinstance(value, bool):
                status_val = "COMPLETED" if value else "NOT_STARTED"
            elif field in ("isrc", "iswc"):
                status_val = "COMPLETED" if (value and str(value).strip()) else "NOT_STARTED"
            else:
                str_val = str(value).strip() if value else ""
                upper_val = str_val.upper()
                if upper_val in ("N/A", "NA", "NOT_APPLICABLE"):
                    status_val = "NOT_APPLICABLE"
                elif upper_val in ("YES", "TRUE", "1"):
                    status_val = "COMPLETED"
                elif str_val:
                    try:
                        float(str_val)
                        status_val = "COMPLETED"
                    except ValueError:
                        status_val = "NOT_STARTED"
                else:
                    status_val = "NOT_STARTED"
            
            db.add(SongChecklistStatus(
                song_id=song.id,
                checklist_item_id=item.id,
                status=status_val
            ))
            if status_val in ("COMPLETED", "NOT_APPLICABLE"):
                completed_weight += item.weight
        
        song.status_health_score = round(min((completed_weight / total_weight) * 100, 100.0), 2)
        
        if creator:
            db.add(SongCredit(
                song_id=song.id,
                creator_id=creator.id,
                role="ARTIST",
                share_percentage=song.publishing_percentage or 100
            ))
    
    for invalid in validation["invalid_rows"]:
        errors.append(f"Row {invalid['row_index'] + 1}: {', '.join(invalid['errors'])}")
    
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"CSV import commit failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to save imported songs: {str(e)}")

    # Persist the original Schedule A upload for audit / re-extraction.
    if request.staged_file_id:
        try:
            from ..services import schedule_a_storage
            from ..models.models import ScheduleAImport
            from ..services.audit_service import log_action

            staged_key = schedule_a_storage.find_staged(org_id, request.staged_file_id)
            if staged_key is not None:
                staged_bytes = schedule_a_storage.open_bytes(staged_key) or b""
                size = len(staged_bytes)
                sha = schedule_a_storage.hash_bytes(staged_bytes)
                fallback_name = staged_key.rsplit("/", 1)[-1]
                if "__" in fallback_name:
                    fallback_name = fallback_name.split("__", 1)[1]
                rec = ScheduleAImport(
                    organization_id=org_id,
                    user_id=current_user.id,
                    creator_id=creator.id if creator else None,
                    creator_name=creator.display_name if creator else None,
                    original_filename=request.staged_filename or fallback_name,
                    stored_path=staged_key,
                    file_size=size,
                    mime_type=request.staged_mime,
                    sha256=sha,
                    extraction_method=request.extraction_method,
                    songs_created=songs_created,
                    songs_failed=len(validation["invalid_rows"]) + (len(validation["valid_rows"]) - songs_created),
                    contract_terms=request.contract_terms or {},
                    document_info=request.document_info or {},
                    is_text_paste=bool(request.is_text_paste),
                )
                db.add(rec)
                db.flush()
                final_key = schedule_a_storage.promote_staged(org_id, staged_key, rec.id)
                rec.stored_path = final_key
                log_action(
                    db,
                    organization_id=org_id,
                    user_id=current_user.id,
                    action="SCHEDULE_A_IMPORTED",
                    entity_type="schedule_a_import",
                    entity_id=rec.id,
                    entity_name=rec.original_filename,
                    details={
                        "songs_created": songs_created,
                        "extraction_method": request.extraction_method,
                        "creator_id": creator.id if creator else None,
                        "is_text_paste": bool(request.is_text_paste),
                        "sha256": sha,
                    },
                )
                db.commit()
            else:
                logger.warning(
                    f"Staged file {request.staged_file_id} not found for org {org_id}; "
                    "skipping Schedule A import audit record."
                )
        except Exception as e:
            db.rollback()
            logger.error(f"Failed to persist ScheduleAImport audit record: {e}", exc_info=True)
            # Do NOT fail the import if audit persistence fails.

    return ImportResult(
        songs_created=songs_created,
        songs_failed=len(validation["invalid_rows"]) + (len(validation["valid_rows"]) - songs_created),
        creator_id=creator.id if creator else None,
        errors=errors[:20]
    )


def calculate_health_score(row_data: Dict[str, Any]) -> float:
    score = 0.0
    total_weight = 100
    
    if row_data.get("isrc"):
        score += 20
    if row_data.get("iswc"):
        score += 15
    if row_data.get("title"):
        score += 10
    if row_data.get("primary_artist"):
        score += 10
    if row_data.get("release_date"):
        score += 10
    if row_data.get("label"):
        score += 5
    if row_data.get("publishing_percentage"):
        score += 10
    if row_data.get("master_percentage"):
        score += 10
    if row_data.get("project_title"):
        score += 5
    if row_data.get("recording_code"):
        score += 5
    
    return min(score, 100.0)


def determine_checklist_status(code: str, row_data: Dict[str, Any]) -> str:
    if code == "ISRC" and row_data.get("isrc"):
        return "COMPLETED"
    if code == "ISWC" and row_data.get("iswc"):
        return "COMPLETED"
    if code == "METADATA" and row_data.get("title") and row_data.get("primary_artist"):
        return "COMPLETED"
    if code == "RELEASE_DATE" and row_data.get("release_date"):
        return "COMPLETED"
    return "NOT_STARTED"
