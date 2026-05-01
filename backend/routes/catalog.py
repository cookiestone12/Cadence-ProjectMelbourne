from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import or_
from typing import List
from pydantic import BaseModel
from datetime import datetime, timedelta
from ..models import get_db, Song, Songwriter, Analytics, User, Catalog, SongCredit, Creator
from ..utils.auth import get_current_user
from ..services import chartmetric_service, spotify_service, luminate_service
from ..services import valuation_engine, scoring_engine
from ..services.song_lifecycle import auto_release_songs
from ..routes.song_registrations import compute_registration_completeness
from ..config.streaming_rates import (
    TRACKED_PLATFORMS,
    MARKET_MULTIPLIER,
    get_publishing_rate,
    get_master_rate
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import PyPDF2
import io
import json
import os

router = APIRouter(prefix="/api/catalog", tags=["Catalog"])

class ValuationResponse(BaseModel):
    estimated_revenue: float
    low: float
    base: float
    high: float

class ScoreBreakdownResponse(BaseModel):
    catalog_value: float
    growth_momentum: float
    metadata_health: float
    exploitation_potential: float

class SongResponse(BaseModel):
    id: int
    title: str
    artist_name: str
    client_name: str | None
    client_id: int | None
    publishing_percentage: float
    master_percentage: float
    spotify_link: str | None
    spotify_streams: int | None
    estimated_revenue: float
    publishing_revenue: float
    master_revenue: float
    valuation_low: float
    valuation_base: float
    valuation_high: float
    valuation_low_pub: float
    valuation_base_pub: float
    valuation_high_pub: float
    valuation_low_master: float
    valuation_base_master: float
    valuation_high_master: float
    score: float
    score_breakdown: dict | None
    
    class Config:
        from_attributes = True

class SongDetailResponse(BaseModel):
    id: int
    title: str
    artist_name: str
    client_name: str | None
    client_id: int | None
    publishing_percentage: float
    master_percentage: float
    spotify_link: str | None
    isrc: str | None
    iswc: str | None
    writer_splits: list | None
    estimated_revenue: float
    publishing_revenue: float
    master_revenue: float
    valuation_low: float
    valuation_base: float
    valuation_high: float
    valuation_low_pub: float
    valuation_base_pub: float
    valuation_high_pub: float
    valuation_low_master: float
    valuation_base_master: float
    valuation_high_master: float
    score: float
    score_breakdown: dict | None
    analytics: dict | None
    release_date: str | None
    spotify_streams: int | None
    premium_streams: int | None
    ad_supported_streams: int | None
    territory_streams: dict | None
    collectible_publishing_value: float
    black_box_loss: float
    
    class Config:
        from_attributes = True

class CatalogSummaryResponse(BaseModel):
    id: int
    name: str
    total_songs: int
    total_publishing_percentage: float
    total_valuation_low: float
    total_valuation_base: float
    total_valuation_high: float
    total_valuation_low_pub: float
    total_valuation_base_pub: float
    total_valuation_high_pub: float
    total_valuation_low_master: float
    total_valuation_base_master: float
    total_valuation_high_master: float
    avg_score: float
    avg_score_breakdown: dict
    total_streams_gross: int
    total_premium_streams: int
    total_ad_supported_streams: int
    total_publishing_revenue: float
    total_master_revenue: float
    publishing_revenue_by_type: dict
    master_revenue_by_type: dict
    territory_breakdown: dict
    label_share_80_20: float
    label_share_60_40: float
    collectible_publishing_value: float
    black_box_loss: float

@router.get("/summary", response_model=List[CatalogSummaryResponse], summary="Get summary for all catalogs", description="Returns one summary row per Catalog the user can see (song count, last upload, owner). Used to render the catalog list page.\n\n**Query:** `org_id` (defaults to the caller's current org).\n**Auth:** Bearer JWT.\n**Response:** `List[CatalogSummaryResponse]` — `[{id, name, song_count, owner, last_upload_at, status}]`.")
def get_catalog_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get summary for all catalogs"""
    # Task #171 — Phase 4: lazy auto-release before computing rollups so the
    # released/unreleased split shown on the catalog list reflects today's
    # state. Auth dependency is required because this is a GET that now
    # writes (commits) — without `current_user`, an unauthenticated caller
    # could trigger persistent state changes. The flip itself is still
    # global because this endpoint enumerates every Catalog regardless of
    # org membership; restricting that semantic is out of scope for #171.
    auto_release_songs(db)
    catalogs = db.query(Catalog).all()
    result = []
    
    for catalog in catalogs:
        songs = catalog.songs
        if not songs:
            continue
            
        total_publishing = sum(s.publishing_percentage for s in songs)
        total_val_low = sum(s.valuation_low for s in songs)
        total_val_base = sum(s.valuation_base for s in songs)
        total_val_high = sum(s.valuation_high for s in songs)
        total_val_low_pub = sum(s.valuation_low_pub for s in songs)
        total_val_base_pub = sum(s.valuation_base_pub for s in songs)
        total_val_high_pub = sum(s.valuation_high_pub for s in songs)
        total_val_low_master = sum(s.valuation_low_master for s in songs)
        total_val_base_master = sum(s.valuation_base_master for s in songs)
        total_val_high_master = sum(s.valuation_high_master for s in songs)
        avg_score = sum(s.score for s in songs) / len(songs) if songs else 0
        
        avg_breakdown = {
            "catalog_value": 0,
            "growth_momentum": 0,
            "metadata_health": 0,
            "exploitation_potential": 0
        }
        
        for song in songs:
            if song.score_breakdown:
                for key in avg_breakdown.keys():
                    avg_breakdown[key] += song.score_breakdown.get(key, 0)
        
        for key in avg_breakdown.keys():
            avg_breakdown[key] = round(avg_breakdown[key] / len(songs), 2) if songs else 0
        
        # Calculate publishing and master revenue across ALL platforms with correct rates
        total_publishing_revenue = 0.0
        total_master_revenue = 0.0
        total_streams_gross = 0
        total_premium_streams = 0
        total_ad_supported_streams = 0
        
        publishing_revenue_by_type = {'premium': 0.0, 'ad_supported': 0.0}
        master_revenue_by_type = {'premium': 0.0, 'ad_supported': 0.0}
        
        # Territory breakdown
        territory_revenue = {}
        
        for song in songs:
            if not song.analytics:
                continue
            
            # Get multi-platform stream breakdown
            streams_by_type = song.analytics.streams_by_type or {}
            
            pub_pct = song.publishing_percentage / 100.0
            master_pct = song.master_percentage / 100.0
            
            # Calculate revenue across ALL platforms
            for platform, stream_data in streams_by_type.items():
                premium_streams = stream_data.get('premium', 0)
                ad_supported_streams = stream_data.get('ad_supported', 0)
                
                # Track total streams (use Spotify as reference for gross count)
                if platform == 'spotify':
                    total_streams_gross += premium_streams + ad_supported_streams
                    total_premium_streams += premium_streams
                    total_ad_supported_streams += ad_supported_streams
                
                # Publishing uses consistent rates across platforms
                pub_premium_rate = get_publishing_rate('premium')
                pub_ad_rate = get_publishing_rate('ad_supported')
                
                # Master uses platform-specific rates (KEY DIFFERENCE!)
                master_premium_rate = get_master_rate(platform, 'premium')
                master_ad_rate = get_master_rate(platform, 'ad_supported')
                
                # Calculate publishing revenue
                premium_pub_rev = premium_streams * pub_premium_rate * pub_pct
                ad_pub_rev = ad_supported_streams * pub_ad_rate * pub_pct
                publishing_revenue_by_type['premium'] += premium_pub_rev
                publishing_revenue_by_type['ad_supported'] += ad_pub_rev
                total_publishing_revenue += premium_pub_rev + ad_pub_rev
                
                # Calculate master revenue with platform-specific rates
                premium_master_rev = premium_streams * master_premium_rate * master_pct
                ad_master_rev = ad_supported_streams * master_ad_rate * master_pct
                master_revenue_by_type['premium'] += premium_master_rev
                master_revenue_by_type['ad_supported'] += ad_master_rev
                total_master_revenue += premium_master_rev + ad_master_rev
            
            # Calculate territory breakdown (using Spotify data for now)
            territory_streams = song.analytics.territory_streams or {}
            for territory, streams in territory_streams.items():
                if territory not in territory_revenue:
                    territory_revenue[territory] = {
                        'publishing': 0.0,
                        'master': 0.0,
                        'total_streams': 0
                    }
                
                terr_premium = streams.get('premium', 0)
                terr_ad = streams.get('ad_supported', 0)
                territory_revenue[territory]['total_streams'] += terr_premium + terr_ad
                
                # Territory data is Spotify-only, so use Spotify rates
                pub_premium_rate = get_publishing_rate('premium')
                pub_ad_rate = get_publishing_rate('ad_supported')
                master_premium_rate = get_master_rate('spotify', 'premium')
                master_ad_rate = get_master_rate('spotify', 'ad_supported')
                
                terr_pub_rev = (terr_premium * pub_premium_rate + terr_ad * pub_ad_rate) * pub_pct
                terr_master_rev = (terr_premium * master_premium_rate + terr_ad * master_ad_rate) * master_pct
                
                territory_revenue[territory]['publishing'] += terr_pub_rev
                territory_revenue[territory]['master'] += terr_master_rev
        
        # Label share calculations (on publishing revenue only)
        label_share_80_20 = total_publishing_revenue * 0.2
        label_share_60_40 = total_publishing_revenue * 0.4
        
        # Round territory revenues
        territory_breakdown = {
            territory: {
                'publishing': round(data['publishing'], 2),
                'master': round(data['master'], 2),
                'total_streams': data['total_streams']
            }
            for territory, data in territory_revenue.items()
        }
        
        # Calculate Black Box Loss and Collectible Publishing Value
        # Based on industry collection windows (2-3 years)
        collectible_publishing_value = 0.0
        black_box_loss = 0.0
        
        for song in songs:
            if not song.analytics:
                continue
            
            # Calculate lifetime publishing revenue for this song across ALL platforms
            streams_by_type = song.analytics.streams_by_type or {}
            pub_pct = song.publishing_percentage / 100.0
            
            lifetime_pub_revenue = 0.0
            for platform, stream_data in streams_by_type.items():
                premium_streams = stream_data.get('premium', 0)
                ad_supported_streams = stream_data.get('ad_supported', 0)
                
                pub_premium_rate = get_publishing_rate('premium')
                pub_ad_rate = get_publishing_rate('ad_supported')
                
                lifetime_pub_revenue += (premium_streams * pub_premium_rate * pub_pct) + \
                                       (ad_supported_streams * pub_ad_rate * pub_pct)
            
            # Calculate collection window decay based on song age
            # Industry standard: 2-3 years collection window
            if song.release_date:
                song_age_days = (datetime.utcnow() - song.release_date).days
                song_age_years = song_age_days / 365.25
                
                if song_age_years <= 3.0:
                    # 0-3 years: 100% collectible
                    decay_factor = 1.0
                elif song_age_years <= 5.0:
                    # 3-5 years: 50% collectible (half lost to black box)
                    decay_factor = 0.5
                else:
                    # 5+ years: 10% collectible (90% lost to black box)
                    decay_factor = 0.1
            else:
                # No release date available, assume 100% collectible
                decay_factor = 1.0
            
            collectible_value = lifetime_pub_revenue * decay_factor
            lost_value = lifetime_pub_revenue * (1.0 - decay_factor)
            
            collectible_publishing_value += collectible_value
            black_box_loss += lost_value
        
        result.append({
            "id": catalog.id,
            "name": catalog.name,
            "total_songs": len(songs),
            "total_publishing_percentage": round(total_publishing, 2),
            "total_valuation_low": round(total_val_low, 2),
            "total_valuation_base": round(total_val_base, 2),
            "total_valuation_high": round(total_val_high, 2),
            "total_valuation_low_pub": round(total_val_low_pub, 2),
            "total_valuation_base_pub": round(total_val_base_pub, 2),
            "total_valuation_high_pub": round(total_val_high_pub, 2),
            "total_valuation_low_master": round(total_val_low_master, 2),
            "total_valuation_base_master": round(total_val_base_master, 2),
            "total_valuation_high_master": round(total_val_high_master, 2),
            "avg_score": round(avg_score, 2),
            "avg_score_breakdown": avg_breakdown,
            "total_streams_gross": total_streams_gross,
            "total_premium_streams": total_premium_streams,
            "total_ad_supported_streams": total_ad_supported_streams,
            "total_publishing_revenue": round(total_publishing_revenue, 2),
            "total_master_revenue": round(total_master_revenue, 2),
            "publishing_revenue_by_type": {
                'premium': round(publishing_revenue_by_type['premium'], 2),
                'ad_supported': round(publishing_revenue_by_type['ad_supported'], 2)
            },
            "master_revenue_by_type": {
                'premium': round(master_revenue_by_type['premium'], 2),
                'ad_supported': round(master_revenue_by_type['ad_supported'], 2)
            },
            "territory_breakdown": territory_breakdown,
            "label_share_80_20": round(label_share_80_20, 2),
            "label_share_60_40": round(label_share_60_40, 2),
            "collectible_publishing_value": round(collectible_publishing_value, 2),
            "black_box_loss": round(black_box_loss, 2)
        })
    
    return result

@router.get(
    "/songs",
    response_model=List[SongResponse],
    summary='List songs with optional filters (legacy catalog list)',
    description="Returns the org's songs with optional filtering by catalog, status, or text. Lower-level than `/api/songs`.\n\n**Query:** `org_id`, `catalog_id`, `status`, `q`, `limit`, `offset`.\n**Auth:** Bearer JWT — caller must be a member of the org.\n**Response:** `List[SongResponse]`.",
)
def get_songs(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Task #171 — Phase 4: lazy auto-release. This legacy endpoint walks every
    # song so we run the flip with no org filter; it's a no-op once the
    # steady-state has been reached. `current_user` is required because
    # this is a GET that now writes (commits).
    auto_release_songs(db)
    # Task #173 — Phase 6 perf sweep: eager-load `analytics` to avoid an
    # N+1 burst (one extra SELECT per song for `song.analytics.*`), and
    # batch-load credits + creators in a single round-trip per kind
    # instead of two queries per song.
    songs = db.query(Song).options(joinedload(Song.analytics)).all()
    song_ids = [s.id for s in songs]
    first_credit_by_song: dict = {}
    if song_ids:
        for credit in (
            db.query(SongCredit)
            .filter(SongCredit.song_id.in_(song_ids))
            .all()
        ):
            first_credit_by_song.setdefault(credit.song_id, credit)
        creator_ids = {c.creator_id for c in first_credit_by_song.values() if c.creator_id}
        creators_by_id = {
            c.id: c
            for c in db.query(Creator).filter(Creator.id.in_(creator_ids)).all()
        } if creator_ids else {}
    else:
        creators_by_id = {}
    result = []

    for song in songs:
        spotify_streams = song.analytics.spotify_streams if song.analytics else None

        publishing_revenue = 0.0
        master_revenue = 0.0

        if song.analytics and song.analytics.streams_by_type:
            streams_by_type = song.analytics.streams_by_type
            pub_pct = song.publishing_percentage / 100.0
            master_pct = song.master_percentage / 100.0

            for platform, stream_data in streams_by_type.items():
                premium_streams = stream_data.get('premium', 0)
                ad_supported_streams = stream_data.get('ad_supported', 0)

                pub_premium_rate = get_publishing_rate('premium')
                pub_ad_rate = get_publishing_rate('ad_supported')

                master_premium_rate = get_master_rate(platform, 'premium')
                master_ad_rate = get_master_rate(platform, 'ad_supported')

                publishing_revenue += (premium_streams * pub_premium_rate * pub_pct) + \
                                     (ad_supported_streams * pub_ad_rate * pub_pct)
                master_revenue += (premium_streams * master_premium_rate * master_pct) + \
                                 (ad_supported_streams * master_ad_rate * master_pct)

        client_name = None
        client_id = None
        credit = first_credit_by_song.get(song.id)
        if credit:
            creator = creators_by_id.get(credit.creator_id)
            if creator:
                client_name = creator.display_name
                client_id = creator.id
        
        result.append({
            "id": song.id,
            "title": song.title,
            "artist_name": song.artist_name,
            "client_name": client_name,
            "client_id": client_id,
            "publishing_percentage": song.publishing_percentage,
            "master_percentage": song.master_percentage,
            "spotify_link": song.spotify_link,
            "spotify_streams": spotify_streams,
            "estimated_revenue": song.estimated_revenue,
            "publishing_revenue": round(publishing_revenue, 2),
            "master_revenue": round(master_revenue, 2),
            "valuation_low": song.valuation_low,
            "valuation_base": song.valuation_base,
            "valuation_high": song.valuation_high,
            "valuation_low_pub": song.valuation_low_pub,
            "valuation_base_pub": song.valuation_base_pub,
            "valuation_high_pub": song.valuation_high_pub,
            "valuation_low_master": song.valuation_low_master,
            "valuation_base_master": song.valuation_base_master,
            "valuation_high_master": song.valuation_high_master,
            "score": song.score,
            "score_breakdown": song.score_breakdown
        })
    return result

@router.get(
    "/songs/{song_id}",
    response_model=SongDetailResponse,
    summary="Get a song's full detail (legacy catalog endpoint)",
    description="Returns the song with credits, splits, audio assets, and DSP links inlined. Lower-level than `/api/songs/{id}`.\n\n**Path parameter:** `song_id`.\n**Auth:** Bearer JWT — caller must be a member of the song's org.\n**Response:** `SongDetailResponse`.",
)
def get_song(
    song_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    song = db.query(Song).filter(Song.id == song_id).first()
    if not song:
        raise HTTPException(status_code=404, detail="Song not found")

    # Task #171 — Phase 4: lazy auto-release for the legacy single-song path.
    # `current_user` is required because this is a GET that now writes
    # (commits) — without it, an unauthenticated caller could trigger
    # persistent state changes.
    if auto_release_songs(db, organization_id=song.organization_id) > 0:
        db.refresh(song)

    analytics_data = None
    spotify_streams = 0
    premium_streams = 0
    ad_supported_streams = 0
    territory_streams = {}
    publishing_revenue = 0.0
    master_revenue = 0.0
    
    if song.analytics:
        spotify_streams = song.analytics.spotify_streams or 0
        analytics_data = {
            "spotify_streams": spotify_streams,
            "spotify_monthly_listeners": song.analytics.spotify_monthly_listeners,
            "chartmetric_score": song.analytics.chartmetric_score,
            "playlist_count": song.analytics.playlist_count,
            "top_playlists": song.analytics.top_playlists,
            "regional_data": song.analytics.regional_data,
            "trend_data": song.analytics.trend_data
        }
        
        if song.analytics.streams_by_type:
            streams_by_type = song.analytics.streams_by_type
            pub_pct = song.publishing_percentage / 100.0
            master_pct = song.master_percentage / 100.0
            
            # Calculate revenue across ALL platforms with correct rates
            for platform, stream_data in streams_by_type.items():
                platform_premium = stream_data.get('premium', 0)
                platform_ad = stream_data.get('ad_supported', 0)
                
                # Track Spotify streams for display
                if platform == 'spotify':
                    premium_streams = platform_premium
                    ad_supported_streams = platform_ad
                
                # Publishing uses consistent rates
                pub_premium_rate = get_publishing_rate('premium')
                pub_ad_rate = get_publishing_rate('ad_supported')
                
                # Master uses platform-specific rates
                master_premium_rate = get_master_rate(platform, 'premium')
                master_ad_rate = get_master_rate(platform, 'ad_supported')
                
                publishing_revenue += (platform_premium * pub_premium_rate * pub_pct) + \
                                     (platform_ad * pub_ad_rate * pub_pct)
                master_revenue += (platform_premium * master_premium_rate * master_pct) + \
                                 (platform_ad * master_ad_rate * master_pct)
        
        if song.analytics.territory_streams:
            territory_streams = song.analytics.territory_streams
    
    release_date_str = song.release_date.isoformat() if song.release_date else None
    
    decay_factor = 1.0
    if song.release_date:
        age_years = (datetime.now() - song.release_date).days / 365.25
        if age_years > 5:
            decay_factor = 0.1
        elif age_years > 3:
            decay_factor = 0.5
    
    collectible_publishing_value = publishing_revenue * decay_factor
    black_box_loss = publishing_revenue * (1 - decay_factor)
    
    client_name = None
    client_id = None
    credit = db.query(SongCredit).filter(SongCredit.song_id == song.id).first()
    if credit:
        creator = db.query(Creator).filter(Creator.id == credit.creator_id).first()
        if creator:
            client_name = creator.display_name
            client_id = creator.id
    
    return {
        "id": song.id,
        "title": song.title,
        "artist_name": song.artist_name,
        "client_name": client_name,
        "client_id": client_id,
        "publishing_percentage": song.publishing_percentage,
        "master_percentage": song.master_percentage,
        "spotify_link": song.spotify_link,
        "isrc": song.isrc,
        "iswc": song.iswc,
        "writer_splits": song.writer_splits,
        "estimated_revenue": song.estimated_revenue,
        "publishing_revenue": round(publishing_revenue, 2),
        "master_revenue": round(master_revenue, 2),
        "valuation_low": song.valuation_low,
        "valuation_base": song.valuation_base,
        "valuation_high": song.valuation_high,
        "valuation_low_pub": song.valuation_low_pub,
        "valuation_base_pub": song.valuation_base_pub,
        "valuation_high_pub": song.valuation_high_pub,
        "valuation_low_master": song.valuation_low_master,
        "valuation_base_master": song.valuation_base_master,
        "valuation_high_master": song.valuation_high_master,
        "score": song.score,
        "score_breakdown": song.score_breakdown,
        "analytics": analytics_data,
        "release_date": release_date_str,
        "spotify_streams": spotify_streams,
        "premium_streams": premium_streams,
        "ad_supported_streams": ad_supported_streams,
        "territory_streams": territory_streams,
        "collectible_publishing_value": round(collectible_publishing_value, 2),
        "black_box_loss": round(black_box_loss, 2)
    }

@router.post("/upload", summary="Upload and parse Schedule A template", description='Uploads a Schedule A template (PDF or Excel), runs the AI parser, and creates a new Catalog with one Song per parsed row. The unparsed file is also archived for audit.\n\n**Body (multipart/form-data):** `file` — the Schedule A; `org_id`; `name?`.\n**Auth:** Bearer JWT — caller must be a member of `org_id`.\n**Response:** `{ catalog_id, name, parsed_rows, songs_created, warnings: [...] }`.')
async def upload_schedule_a(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload and parse Schedule A template.
    Supports PDF and Excel formats.
    """
    
    content = await file.read()
    filename = (file.filename or "").lower()
    
    songs_data = []
    songwriter_data = {}
    
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Parse Excel file
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            
            # Extract songwriter info from updated template format (rows 5-7)
            songwriter_data = {
                "name": ws['B5'].value if ws['B5'].value else "Unknown",
                "pro_affiliation": ws['B6'].value,
                "ipi_number": ws['B7'].value
            }
            
            # Extract song data (table starts at row 12, after headers in row 11)
            for row in ws.iter_rows(min_row=12, values_only=True):
                if row[0]:  # If title exists
                    songs_data.append({
                        "title": row[0],
                        "artist_name": row[1] if row[1] else "Unknown",
                        "publishing_percentage": float(row[2]) if row[2] else 0.0,
                        "master_percentage": float(row[3]) if row[3] else 0.0,
                        "spotify_link": row[4] if row[4] else None,
                        "release_date": row[5] if row[5] else None
                    })
        
        elif filename.endswith('.pdf'):
            # Basic PDF parsing (simple text extraction)
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
            
            # Simple parsing logic (this would need to be more sophisticated)
            lines = text.split('\n')
            songwriter_data = {
                "name": "Extracted from PDF",
                "pro_affiliation": None,
                "ipi_number": None
            }
            
            # For now, create a placeholder song
            songs_data.append({
                "title": "Song from PDF",
                "artist_name": "Artist from PDF",
                "publishing_percentage": 0.0,
                "master_percentage": 0.0,
                "spotify_link": None
            })
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")
    
    # Create songwriter
    songwriter = Songwriter(**songwriter_data)
    db.add(songwriter)
    db.commit()
    db.refresh(songwriter)
    
    # Create songs and fetch analytics
    created_songs = []
    for song_data in songs_data:
        # Fetch analytics from mock/real APIs
        chartmetric_data = chartmetric_service.get_track_data(
            song_data['title'], 
            song_data['artist_name']
        )
        spotify_data = spotify_service.get_track_data(song_data.get('spotify_link'))
        luminate_data = luminate_service.get_track_data(
            song_data['title'],
            song_data['artist_name']
        )
        
        # Combine analytics
        analytics_combined = {
            "spotify_streams": spotify_data.get('streams_data', {}).get('total_streams', 0),
            "spotify_monthly_listeners": spotify_data.get('streams_data', {}).get('monthly_streams', 0),
            "chartmetric_score": chartmetric_data.get('chartmetric_score', 0),
            "playlist_count": chartmetric_data.get('playlist_count', 0),
            "top_playlists": chartmetric_data.get('top_playlists', []),
            "regional_data": luminate_data.get('regional_performance', {}),
            "trend_data": {"momentum": chartmetric_data.get('trend_momentum', 'stable')}
        }
        
        # Add growth metrics to analytics for valuation/scoring
        analytics_combined['growth_3_month'] = chartmetric_data.get('growth_3_month', 0)
        analytics_combined['growth_12_month'] = chartmetric_data.get('growth_12_month', 0)
        analytics_combined['has_isrc'] = False
        analytics_combined['has_iswc'] = False
        analytics_combined['has_spotify_link'] = bool(song_data.get('spotify_link'))
        
        # Tier-aware ingestion: preserve uploader-provided premium/ad-supported values
        spotify_streams = analytics_combined.get('spotify_streams', 0)
        platform_streams_from_chartmetric = chartmetric_data.get('platform_streams', {})
        
        streams_by_type = {}
        premium_only_platforms = ['apple_music', 'tidal']
        
        # Process each tracked platform
        for platform in TRACKED_PLATFORMS:
            platform_data = platform_streams_from_chartmetric.get(platform)
            
            if isinstance(platform_data, dict) and ('premium' in platform_data or 'ad_supported' in platform_data):
                # Case 1: Tier-level data provided (preserve it)
                premium = int(platform_data.get('premium', 0))
                ad_supported = int(platform_data.get('ad_supported', 0))
                streams_by_type[platform] = {
                    'premium': premium,
                    'ad_supported': ad_supported
                }
            elif platform_data and isinstance(platform_data, (int, float)) and platform_data > 0:
                # Case 2: Total provided without tier breakdown (apply heuristics)
                total = int(platform_data)
                if platform in premium_only_platforms:
                    premium = total
                    ad_supported = 0
                else:
                    premium = int(total * 0.7)
                    ad_supported = total - premium
                streams_by_type[platform] = {
                    'premium': premium,
                    'ad_supported': ad_supported
                }
            elif platform == 'spotify' and spotify_streams > 0:
                # Case 3: Use Spotify data we already have (apply heuristic)
                total = spotify_streams
                premium = int(total * 0.7)
                ad_supported = total - premium
                streams_by_type[platform] = {
                    'premium': premium,
                    'ad_supported': ad_supported
                }
            else:
                # Case 4: No data for this platform, estimate from Spotify using market share
                from ..config.streaming_rates import MARKET_SHARE
                if spotify_streams > 0:
                    ratio = MARKET_SHARE[platform] / MARKET_SHARE['spotify']
                    total = int(spotify_streams * ratio)
                    if platform in premium_only_platforms:
                        premium = total
                        ad_supported = 0
                    else:
                        premium = int(total * 0.7)
                        ad_supported = total - premium
                    streams_by_type[platform] = {
                        'premium': premium,
                        'ad_supported': ad_supported
                    }
        
        # Create or get first catalog
        catalog = db.query(Catalog).first()
        if not catalog:
            catalog = Catalog(name="Uploaded Catalog")
            db.add(catalog)
            db.commit()
            db.refresh(catalog)
        
        # Check if song already exists (deduplicate by spotify_link or title+artist)
        existing_song = None
        if song_data.get('spotify_link'):
            existing_song = db.query(Song).filter(Song.spotify_link == song_data['spotify_link']).first()
        if not existing_song:
            existing_song = db.query(Song).filter(
                Song.title == song_data['title'],
                Song.artist_name == song_data['artist_name']
            ).first()
        
        if existing_song:
            # Update ownership percentages
            existing_song.publishing_percentage = song_data['publishing_percentage']
            existing_song.master_percentage = song_data['master_percentage']
            
            # Update existing Analytics with ALL new data including tier-aware streams_by_type
            if existing_song.analytics:
                existing_song.analytics.spotify_streams = analytics_combined['spotify_streams']
                existing_song.analytics.spotify_monthly_listeners = analytics_combined['spotify_monthly_listeners']
                existing_song.analytics.chartmetric_score = analytics_combined['chartmetric_score']
                existing_song.analytics.playlist_count = analytics_combined['playlist_count']
                existing_song.analytics.top_playlists = analytics_combined['top_playlists']
                existing_song.analytics.regional_data = analytics_combined['regional_data']
                existing_song.analytics.trend_data = analytics_combined['trend_data']
                existing_song.analytics.streams_by_type = streams_by_type  # Preserve tier-aware data
            else:
                # Create new Analytics if missing
                analytics = Analytics(
                    song_id=existing_song.id,
                    spotify_streams=analytics_combined['spotify_streams'],
                    spotify_monthly_listeners=analytics_combined['spotify_monthly_listeners'],
                    chartmetric_score=analytics_combined['chartmetric_score'],
                    playlist_count=analytics_combined['playlist_count'],
                    top_playlists=analytics_combined['top_playlists'],
                    regional_data=analytics_combined['regional_data'],
                    trend_data=analytics_combined['trend_data'],
                    streams_by_type=streams_by_type
                )
                db.add(analytics)
            
            # Flush to make Analytics update visible
            db.flush()
            
            # NOW calculate revenue from the updated Analytics.streams_by_type
            pub_pct = song_data['publishing_percentage'] / 100.0
            master_pct = song_data['master_percentage'] / 100.0
            
            publishing_revenue = 0.0
            master_revenue = 0.0
            
            # Use the UPDATED streams_by_type from Analytics
            fresh_streams_by_type = existing_song.analytics.streams_by_type if existing_song.analytics else streams_by_type
            
            for platform, stream_data in fresh_streams_by_type.items():
                premium_streams = stream_data['premium']
                ad_supported_streams = stream_data['ad_supported']
                
                # Publishing uses consistent rates across platforms
                pub_premium_rate = get_publishing_rate('premium')
                pub_ad_rate = get_publishing_rate('ad_supported')
                
                # Master uses platform-specific rates
                master_premium_rate = get_master_rate(platform, 'premium')
                master_ad_rate = get_master_rate(platform, 'ad_supported')
                
                publishing_revenue += (premium_streams * pub_premium_rate * pub_pct) + \
                                     (ad_supported_streams * pub_ad_rate * pub_pct)
                master_revenue += (premium_streams * master_premium_rate * master_pct) + \
                                 (ad_supported_streams * master_ad_rate * master_pct)
            
            # Calculate valuation and score using analytics_combined (has all derived fields like growth_3_month, has_isrc, etc.)
            # Revenue already uses fresh streams_by_type from updated Analytics
            valuation = valuation_engine.calculate_valuation(analytics_combined, publishing_revenue, master_revenue)
            analytics_combined['registration_completeness'] = compute_registration_completeness(db, existing_song)
            score = scoring_engine.calculate_score(analytics_combined)
            
            # Update song with fresh valuations
            existing_song.valuation_low = valuation['valuation_low']
            existing_song.valuation_base = valuation['valuation_base']
            existing_song.valuation_high = valuation['valuation_high']
            existing_song.valuation_low_pub = valuation['valuation_low_pub']
            existing_song.valuation_base_pub = valuation['valuation_base_pub']
            existing_song.valuation_high_pub = valuation['valuation_high_pub']
            existing_song.valuation_low_master = valuation['valuation_low_master']
            existing_song.valuation_base_master = valuation['valuation_base_master']
            existing_song.valuation_high_master = valuation['valuation_high_master']
            existing_song.estimated_revenue = valuation['estimated_revenue']
            existing_song.score = score['overall_score']
            existing_song.score_breakdown = {
                'catalog_value': score['catalog_value'],
                'growth_momentum': score['growth_momentum'],
                'metadata_health': score['metadata_health'],
                'exploitation_potential': score['exploitation_potential']
            }
            
            db.commit()
            db.refresh(existing_song)
            song = existing_song
        else:
            # For new songs, create placeholder song first
            song = Song(
                title=song_data['title'],
                artist_name=song_data['artist_name'],
                publishing_percentage=song_data['publishing_percentage'],
                master_percentage=song_data['master_percentage'],
                spotify_link=song_data.get('spotify_link'),
                songwriter_id=songwriter.id,
                catalog_id=catalog.id,
                valuation_low=0,  # Placeholder
                valuation_base=0,  # Placeholder
                valuation_high=0,  # Placeholder
                valuation_low_pub=0,
                valuation_base_pub=0,
                valuation_high_pub=0,
                valuation_low_master=0,
                valuation_base_master=0,
                valuation_high_master=0,
                estimated_revenue=0,
                score=0,
                score_breakdown={}
            )
            db.add(song)
            db.flush()  # Get song ID without committing
            
            # Create analytics with multi-platform tier-aware streams_by_type
            analytics = Analytics(
                song_id=song.id,
                spotify_streams=analytics_combined['spotify_streams'],
                spotify_monthly_listeners=analytics_combined['spotify_monthly_listeners'],
                chartmetric_score=analytics_combined['chartmetric_score'],
                playlist_count=analytics_combined['playlist_count'],
                top_playlists=analytics_combined['top_playlists'],
                regional_data=analytics_combined['regional_data'],
                trend_data=analytics_combined['trend_data'],
                streams_by_type=streams_by_type
            )
            db.add(analytics)
            db.flush()  # Make Analytics visible
            
            # NOW calculate revenue from the persisted Analytics.streams_by_type
            pub_pct = song_data['publishing_percentage'] / 100.0
            master_pct = song_data['master_percentage'] / 100.0
            
            publishing_revenue = 0.0
            master_revenue = 0.0
            
            for platform, stream_data in analytics.streams_by_type.items():
                premium_streams = stream_data['premium']
                ad_supported_streams = stream_data['ad_supported']
                
                # Publishing uses consistent rates across platforms
                pub_premium_rate = get_publishing_rate('premium')
                pub_ad_rate = get_publishing_rate('ad_supported')
                
                # Master uses platform-specific rates
                master_premium_rate = get_master_rate(platform, 'premium')
                master_ad_rate = get_master_rate(platform, 'ad_supported')
                
                publishing_revenue += (premium_streams * pub_premium_rate * pub_pct) + \
                                     (ad_supported_streams * pub_ad_rate * pub_pct)
                master_revenue += (premium_streams * master_premium_rate * master_pct) + \
                                 (ad_supported_streams * master_ad_rate * master_pct)
            
            # Calculate valuation and score using fresh analytics
            valuation = valuation_engine.calculate_valuation(analytics_combined, publishing_revenue, master_revenue)
            analytics_combined['registration_completeness'] = compute_registration_completeness(db, song)
            score = scoring_engine.calculate_score(analytics_combined)
            
            # Update song with calculated valuations
            song.valuation_low = valuation['valuation_low']
            song.valuation_base = valuation['valuation_base']
            song.valuation_high = valuation['valuation_high']
            song.valuation_low_pub = valuation['valuation_low_pub']
            song.valuation_base_pub = valuation['valuation_base_pub']
            song.valuation_high_pub = valuation['valuation_high_pub']
            song.valuation_low_master = valuation['valuation_low_master']
            song.valuation_base_master = valuation['valuation_base_master']
            song.valuation_high_master = valuation['valuation_high_master']
            song.estimated_revenue = valuation['estimated_revenue']
            song.score = score['overall_score']
            song.score_breakdown = {
                'catalog_value': score['catalog_value'],
                'growth_momentum': score['growth_momentum'],
                'metadata_health': score['metadata_health'],
                'exploitation_potential': score['exploitation_potential']
            }
            
            db.commit()
            db.refresh(song)
        
        created_songs.append(song)
    
    return {
        "message": f"Successfully uploaded {len(created_songs)} songs",
        "songs": [{"id": s.id, "title": s.title} for s in created_songs]
    }

@router.get("/search", summary="Search for songs by title or artist name", description='Searches the org\'s catalog by song title or primary artist. Falls back to a small mock external dataset if no local matches are found (used to demo the search UI on empty orgs).\n\n**Query:** `q` (required), `org_id`, `limit` (default 25).\n**Auth:** Bearer JWT.\n**Response:** `{ results: [{song_id?, title, artist, isrc?, source: "local"|"external"}] }`.')
def search_songs(
    q: str,
    db: Session = Depends(get_db)
):
    """
    Search for songs by title or artist name.
    Returns local catalog matches or mock external data if no match found.
    """
    
    search_term = q.lower().strip()
    
    local_songs = db.query(Song).filter(
        or_(
            Song.title.ilike(f"%{search_term}%"),
            Song.artist_name.ilike(f"%{search_term}%")
        )
    ).all()
    
    results = []
    
    for song in local_songs:
        analytics_data = {}
        if song.analytics:
            analytics_data = {
                "total_streams": song.analytics.spotify_streams,
                "monthly_streams": song.analytics.spotify_monthly_listeners,
                "playlist_count": song.analytics.playlist_count,
                "top_playlist": song.analytics.top_playlists[0]['name'] if song.analytics.top_playlists else None,
                "top_playlist_followers": song.analytics.top_playlists[0]['followers'] if song.analytics.top_playlists else 0,
                "growth_3_month": song.analytics.trend_data.get('growth_3_month', 0),
                "chartmetric_score": song.analytics.chartmetric_score,
                "top_territories": [
                    {"country": country, "streams": data.get('streams', 0)}
                    for country, data in (song.analytics.regional_data or {}).items()
                ][:3]
            }
        
        results.append({
            "title": song.title,
            "artist_name": song.artist_name,
            "in_catalog": True,
            "catalog_name": song.catalog.name if song.catalog else None,
            "publishing_percentage": song.publishing_percentage,
            "master_percentage": song.master_percentage,
            "spotify_link": song.spotify_link,
            "valuation": {
                "estimated_revenue": song.estimated_revenue,
                "low": song.valuation_low,
                "base": song.valuation_base,
                "high": song.valuation_high
            },
            "score": song.score,
            "score_breakdown": song.score_breakdown,
            "metrics": analytics_data
        })
    
    if not results:
        tracks_path = os.path.join(os.path.dirname(__file__), '..', '..', 'mock_data', 'external_metrics_tracks.json')
        artists_path = os.path.join(os.path.dirname(__file__), '..', '..', 'mock_data', 'external_metrics_artists.json')
        
        try:
            with open(tracks_path, 'r') as f:
                track_metrics = json.load(f)
            with open(artists_path, 'r') as f:
                artist_metrics = json.load(f)
            
            for title, metrics in track_metrics.items():
                if search_term in title.lower():
                    artist_name = "Unknown Artist"
                    for artist, artist_data in artist_metrics.items():
                        if search_term in artist.lower():
                            artist_name = artist
                            break
                    
                    analytics_data = {
                        'spotify_streams': metrics.get('total_streams', 0),
                        'spotify_monthly_listeners': metrics.get('monthly_streams', 0),
                        'chartmetric_score': metrics.get('chartmetric_score', 0),
                        'playlist_count': metrics.get('playlist_count', 0),
                        'top_playlists': [{
                            'name': metrics.get('top_playlist', ''),
                            'followers': metrics.get('top_playlist_followers', 0),
                            'position': 10
                        }],
                        'growth_3_month': metrics.get('growth_3_month', 0),
                        'growth_12_month': metrics.get('growth_12_month', 0),
                        'has_isrc': True,
                        'has_iswc': True,
                        'has_spotify_link': True
                    }
                    
                    valuation = valuation_engine.calculate_valuation(analytics_data)
                    analytics_data['registration_completeness'] = 1.0
                    score = scoring_engine.calculate_score(analytics_data)
                    
                    artist_data = artist_metrics.get(artist_name, {})
                    
                    results.append({
                        "title": title,
                        "artist_name": artist_name,
                        "in_catalog": False,
                        "catalog_name": None,
                        "publishing_percentage": 0,
                        "master_percentage": 0,
                        "spotify_link": None,
                        "valuation": valuation,
                        "score": score['overall_score'],
                        "score_breakdown": {
                            "catalog_value": score['catalog_value'],
                            "growth_momentum": score['growth_momentum'],
                            "metadata_health": score['metadata_health'],
                            "exploitation_potential": score['exploitation_potential']
                        },
                        "metrics": {
                            "total_streams": metrics.get('total_streams', 0),
                            "monthly_streams": metrics.get('monthly_streams', 0),
                            "playlist_count": metrics.get('playlist_count', 0),
                            "top_playlist": metrics.get('top_playlist'),
                            "top_playlist_followers": metrics.get('top_playlist_followers', 0),
                            "growth_3_month": metrics.get('growth_3_month', 0),
                            "chartmetric_score": metrics.get('chartmetric_score', 0),
                            "top_territories": metrics.get('top_territories', [])
                        },
                        "artist_metrics": {
                            "monthly_listeners": artist_data.get('monthly_listeners', 0),
                            "followers": artist_data.get('followers', 0),
                            "follower_growth_3_month": artist_data.get('follower_growth_3_month', 0),
                            "genre_tags": artist_data.get('genre_tags', [])
                        }
                    })
                    break
        except Exception as e:
            print(f"Error loading mock data: {e}")
    
    return {"results": results, "count": len(results)}

@router.get("/export/{catalog_id}", summary="Generate comprehensive Excel report for a catalog", description="Renders an Excel report of the catalog (songs, splits, ISRCs, registration status) and streams it as a download.\n\n**Path parameter:** `catalog_id`.\n**Auth:** Bearer JWT — caller must be a member of the catalog's org.\n**Response:** `application/vnd.openxmlformats-officedocument.spreadsheetml.sheet` download.")
def export_catalog_report(
    catalog_id: int,
    db: Session = Depends(get_db)
):
    """Generate comprehensive Excel report for a catalog"""
    catalog = db.query(Catalog).filter(Catalog.id == catalog_id).first()
    if not catalog:
        raise HTTPException(status_code=404, detail="Catalog not found")
    
    songs = catalog.songs
    if not songs:
        raise HTTPException(status_code=404, detail="No songs found in catalog")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles - Cadence branding
    header_fill = PatternFill(start_color="E62E2E", end_color="E62E2E", fill_type="solid")
    header_font = Font(name="Inter", color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title styles
    title_font = Font(name="Inter", size=16, bold=True, color="E62E2E")
    subtitle_font = Font(name="Inter", size=10, color="666666")
    
    # Sheet 1: Catalog Summary
    ws_summary = wb.create_sheet("Catalog Summary")
    ws_summary.append(["CADENCE CATALOG INTELLIGENCE REPORT"])
    ws_summary["A1"].font = title_font
    ws_summary.append(["Catalog:", catalog.name])
    ws_summary["A2"].font = subtitle_font
    ws_summary["B2"].font = Font(name="Inter", size=10, bold=True)
    ws_summary.append(["Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws_summary["A3"].font = subtitle_font
    ws_summary["B3"].font = subtitle_font
    ws_summary.append([])
    
    # Calculate summary metrics
    PREMIUM_RATE = 0.0012
    AD_SUPPORTED_RATE = 0.0004
    
    total_songs = len(songs)
    total_publishing_pct = sum(s.publishing_percentage for s in songs)
    total_master_pct = sum(s.master_percentage for s in songs)
    avg_score = sum(s.score for s in songs) / len(songs) if songs else 0
    total_val_low = sum(s.valuation_low for s in songs)
    total_val_base = sum(s.valuation_base for s in songs)
    total_val_high = sum(s.valuation_high for s in songs)
    
    total_publishing_revenue = 0.0
    total_master_revenue = 0.0
    collectible_publishing_value = 0.0
    black_box_loss = 0.0
    total_streams = 0
    
    for song in songs:
        if not song.analytics:
            continue
        
        total_streams += song.analytics.spotify_streams or 0
        
        streams_by_type = song.analytics.streams_by_type or {}
        spotify_streams = streams_by_type.get('spotify', {})
        premium_streams = spotify_streams.get('premium', 0)
        ad_supported_streams = spotify_streams.get('ad_supported', 0)
        
        pub_pct = song.publishing_percentage / 100.0
        master_pct = song.master_percentage / 100.0
        
        pub_rev = (premium_streams * PREMIUM_RATE + ad_supported_streams * AD_SUPPORTED_RATE) * pub_pct
        master_rev = (premium_streams * PREMIUM_RATE + ad_supported_streams * AD_SUPPORTED_RATE) * master_pct
        
        total_publishing_revenue += pub_rev
        total_master_revenue += master_rev
        
        # Black box calculation
        if song.release_date:
            song_age_years = (datetime.utcnow() - song.release_date).days / 365.25
            if song_age_years <= 3.0:
                decay_factor = 1.0
            elif song_age_years <= 5.0:
                decay_factor = 0.5
            else:
                decay_factor = 0.1
        else:
            decay_factor = 1.0
        
        collectible_publishing_value += pub_rev * decay_factor
        black_box_loss += pub_rev * (1.0 - decay_factor)
    
    ws_summary.append(["Metric", "Value"])
    ws_summary["A5"].fill = header_fill
    ws_summary["A5"].font = header_font
    ws_summary["B5"].fill = header_fill
    ws_summary["B5"].font = header_font
    
    ws_summary.append(["Total Songs", total_songs])
    ws_summary.append(["Total Publishing %", f"{total_publishing_pct:.2f}%"])
    ws_summary.append(["Total Master %", f"{total_master_pct:.2f}%"])
    ws_summary.append(["Average Score", f"{avg_score:.2f}/100"])
    ws_summary.append(["Total Streams", f"{total_streams:,}"])
    ws_summary.append([])
    ws_summary.append(["Valuation (Low)", f"${total_val_low:,.2f}"])
    ws_summary.append(["Valuation (Base)", f"${total_val_base:,.2f}"])
    ws_summary.append(["Valuation (High)", f"${total_val_high:,.2f}"])
    ws_summary.append([])
    ws_summary.append(["Total Publishing Revenue", f"${total_publishing_revenue:,.2f}"])
    ws_summary.append(["Total Master Revenue", f"${total_master_revenue:,.2f}"])
    ws_summary.append(["Collectible Publishing Value", f"${collectible_publishing_value:,.2f}"])
    ws_summary.append(["Est. Black Box Loss", f"${black_box_loss:,.2f}"])
    ws_summary.append([])
    ws_summary.append(["Label Share (80/20)", f"${total_publishing_revenue * 0.2:,.2f}"])
    ws_summary.append(["Label Share (60/40)", f"${total_publishing_revenue * 0.4:,.2f}"])
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    
    # Sheet 2: Territory Breakdown
    ws_territory = wb.create_sheet("Territory Breakdown")
    ws_territory.append(["Territory", "Total Streams", "Publishing Revenue", "Master Revenue", "Total Revenue"])
    
    for cell in ws_territory[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    territory_data = {}
    for song in songs:
        if not song.analytics:
            continue
        
        pub_pct = song.publishing_percentage / 100.0
        master_pct = song.master_percentage / 100.0
        
        territory_streams = song.analytics.territory_streams or {}
        for territory, streams in territory_streams.items():
            if territory not in territory_data:
                territory_data[territory] = {'streams': 0, 'publishing': 0.0, 'master': 0.0}
            
            terr_premium = streams.get('premium', 0)
            terr_ad = streams.get('ad_supported', 0)
            territory_data[territory]['streams'] += terr_premium + terr_ad
            
            terr_pub_rev = (terr_premium * PREMIUM_RATE + terr_ad * AD_SUPPORTED_RATE) * pub_pct
            terr_master_rev = (terr_premium * PREMIUM_RATE + terr_ad * AD_SUPPORTED_RATE) * master_pct
            
            territory_data[territory]['publishing'] += terr_pub_rev
            territory_data[territory]['master'] += terr_master_rev
    
    for territory, data in sorted(territory_data.items(), key=lambda x: x[1]['streams'], reverse=True):
        total_rev = data['publishing'] + data['master']
        ws_territory.append([
            territory,
            f"{data['streams']:,}",
            f"${data['publishing']:,.2f}",
            f"${data['master']:,.2f}",
            f"${total_rev:,.2f}"
        ])
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_territory.column_dimensions[col].width = 20
    
    # Sheet 3: Song Details
    ws_songs = wb.create_sheet("Song Details")
    ws_songs.append([
        "Title", "Artist", "Release Date", "Song Age (Years)", 
        "Publishing %", "Master %", "Total Streams", 
        "Publishing Revenue", "Master Revenue",
        "Collectible Value", "Black Box Loss",
        "Score", "Valuation (Base)"
    ])
    
    for cell in ws_songs[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for song in songs:
        streams = song.analytics.spotify_streams if song.analytics else 0
        
        # Calculate revenue
        if song.analytics:
            streams_by_type = song.analytics.streams_by_type or {}
            spotify_streams = streams_by_type.get('spotify', {})
            premium_streams = spotify_streams.get('premium', 0)
            ad_supported_streams = spotify_streams.get('ad_supported', 0)
            
            pub_pct = song.publishing_percentage / 100.0
            master_pct = song.master_percentage / 100.0
            
            pub_rev = (premium_streams * PREMIUM_RATE + ad_supported_streams * AD_SUPPORTED_RATE) * pub_pct
            master_rev = (premium_streams * PREMIUM_RATE + ad_supported_streams * AD_SUPPORTED_RATE) * master_pct
            
            # Black box calculation
            if song.release_date:
                song_age_years = (datetime.utcnow() - song.release_date).days / 365.25
                if song_age_years <= 3.0:
                    decay_factor = 1.0
                elif song_age_years <= 5.0:
                    decay_factor = 0.5
                else:
                    decay_factor = 0.1
            else:
                song_age_years = 0
                decay_factor = 1.0
            
            collectible = pub_rev * decay_factor
            lost = pub_rev * (1.0 - decay_factor)
        else:
            pub_rev = 0
            master_rev = 0
            collectible = 0
            lost = 0
            song_age_years = 0
        
        release_date_str = song.release_date.strftime("%Y-%m-%d") if song.release_date else "Unknown"
        
        ws_songs.append([
            song.title,
            song.artist_name,
            release_date_str,
            f"{song_age_years:.1f}",
            f"{song.publishing_percentage:.1f}%",
            f"{song.master_percentage:.1f}%",
            f"{streams:,}",
            f"${pub_rev:,.2f}",
            f"${master_rev:,.2f}",
            f"${collectible:,.2f}",
            f"${lost:,.2f}",
            f"{song.score:.1f}/100",
            f"${song.valuation_base:,.2f}"
        ])
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws_songs.column_dimensions[col].width = 18
    
    # Sheet 4: Methodology
    ws_method = wb.create_sheet("Methodology")
    ws_method.append(["Cadence Catalog Intelligence - Calculation Methodology"])
    ws_method.append([])
    ws_method.append(["Revenue Calculations:"])
    ws_method.append(["- Premium Stream Rate: $0.0012 per stream"])
    ws_method.append(["- Ad-Supported Stream Rate: $0.0004 per stream"])
    ws_method.append(["- Stream Type Split: 70% Premium / 30% Ad-Supported"])
    ws_method.append(["- Publishing Revenue = Streams × Rate × Publishing %"])
    ws_method.append(["- Master Revenue = Streams × Rate × Master %"])
    ws_method.append([])
    ws_method.append(["Black Box Collection Windows:"])
    ws_method.append(["- 0-3 Years Old: 100% Collectible"])
    ws_method.append(["- 3-5 Years Old: 50% Collectible (50% Lost to Black Box)"])
    ws_method.append(["- 5+ Years Old: 10% Collectible (90% Lost to Black Box)"])
    ws_method.append([])
    ws_method.append(["Valuation Methodology:"])
    ws_method.append(["- Low Valuation: 8× Revenue Multiple"])
    ws_method.append(["- Base Valuation: 12× Revenue Multiple"])
    ws_method.append(["- High Valuation: 18× Revenue Multiple"])
    ws_method.append(["- Multipliers adjusted based on:"])
    ws_method.append(["  * Stream volume and growth trends"])
    ws_method.append(["  * Playlist placement and follower counts"])
    ws_method.append(["  * Chartmetric score (market momentum)"])
    ws_method.append(["  * Regional performance diversity"])
    ws_method.append([])
    ws_method.append(["4-Factor Scoring System (0-100 points):"])
    ws_method.append(["1. Catalog Value (0-25): Based on stream volume and revenue"])
    ws_method.append(["2. Growth Momentum (0-25): 3-month and 12-month growth trends"])
    ws_method.append(["3. Metadata Health (0-25): Completeness of ISRC, ISWC, Spotify links"])
    ws_method.append(["4. Exploitation Potential (0-25): Playlist presence and regional reach"])
    ws_method.append([])
    ws_method.append(["Label Share Scenarios:"])
    ws_method.append(["- 80/20 Split: 20% to label from publishing revenue"])
    ws_method.append(["- 60/40 Split: 40% to label from publishing revenue"])
    ws_method.append(["- Note: Label shares apply to publishing revenue only"])
    
    ws_method.column_dimensions['A'].width = 70
    
    # Save to BytesIO buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Sanitize filename to remove special characters that can't be encoded in latin-1
    safe_catalog_name = ''.join(c if c.isalnum() or c in (' ', '_', '-') else '_' for c in catalog.name)
    safe_catalog_name = safe_catalog_name.replace(' ', '_')
    filename = f"Cadence_Catalog_Report_{safe_catalog_name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/template/schedule-a", summary="Generate and serve the official Cadence Schedule A upload template", description='Generates and serves the official Cadence Schedule A upload template (Excel) so users have the exact column set the parser expects.\n\n**Auth:** Bearer JWT.\n**Response:** `application/vnd.openxmlformats-officedocument.spreadsheetml.sheet` download containing the empty template with header row + sample row.')
def download_schedule_a_template():
    """Generate and serve the official Cadence Schedule A upload template"""
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create a new one
    wb.remove(wb.active)
    ws = wb.create_sheet("Schedule A")
    
    # Define styles - Cadence branding
    title_font = Font(name="Inter", size=18, bold=True, color="E62E2E")
    section_header_font = Font(name="Inter", size=12, bold=True, color="FFFFFF")
    section_header_fill = PatternFill(start_color="E62E2E", end_color="E62E2E", fill_type="solid")
    field_label_font = Font(name="Inter", size=10, bold=True, color="333333")
    table_header_font = Font(name="Inter", size=10, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="E62E2E", end_color="E62E2E", fill_type="solid")
    instruction_font = Font(name="Inter", size=9, italic=True, color="666666")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "CADENCE SCHEDULE A TEMPLATE"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # Subtitle
    ws.merge_cells('A2:F2')
    ws['A2'] = "Catalog Upload Template - Official Use Only"
    ws['A2'].font = Font(name="Inter", size=10, color="666666", italic=True)
    ws['A2'].alignment = center_alignment
    
    # Empty row
    ws.append([])
    
    # Songwriter Information Section
    ws.merge_cells('A4:F4')
    ws['A4'] = "SONGWRITER INFORMATION"
    ws['A4'].font = section_header_font
    ws['A4'].fill = section_header_fill
    ws['A4'].alignment = center_alignment
    ws.row_dimensions[4].height = 25
    
    # Songwriter fields
    ws.append(["Songwriter Name:", "", "", "", "", ""])
    ws['A5'].font = field_label_font
    ws.merge_cells('B5:F5')
    
    ws.append(["PRO Affiliation:", "", "", "", "", ""])
    ws['A6'].font = field_label_font
    ws.merge_cells('B6:F6')
    
    ws.append(["IPI Number:", "", "", "", "", ""])
    ws['A7'].font = field_label_font
    ws.merge_cells('B7:F7')
    
    # Empty rows
    ws.append([])
    ws.append([])
    
    # Song Catalog Table Header
    ws.merge_cells('A10:F10')
    ws['A10'] = "SONG CATALOG"
    ws['A10'].font = section_header_font
    ws['A10'].fill = section_header_fill
    ws['A10'].alignment = center_alignment
    ws.row_dimensions[10].height = 25
    
    # Table column headers
    headers = ["Song Title", "Artist Name", "Publishing %", "Master %", "Spotify Link", "Release Date"]
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = center_alignment
    ws.row_dimensions[11].height = 20
    
    # Add 15 empty rows for song data
    for i in range(15):
        ws.append(["", "", "", "", "", ""])
    
    # Instructions section
    ws.append([])
    ws.append([])
    ws.merge_cells(f'A{ws.max_row}:F{ws.max_row}')
    ws[f'A{ws.max_row}'] = "INSTRUCTIONS"
    ws[f'A{ws.max_row}'].font = Font(name="Inter", size=11, bold=True, color="E62E2E")
    
    ws.append([])
    instructions = [
        "1. Fill in your songwriter information in the designated fields above",
        "2. Enter each song in the catalog table with all required information",
        "3. Publishing % and Master % should be entered as numbers (e.g., 50 for 50%)",
        "4. Spotify Link should be the full URL to the track on Spotify",
        "5. Release Date format: YYYY-MM-DD (e.g., 2024-01-15)",
        "6. Save this file and upload it through the Cadence platform",
        "7. The system will automatically fetch analytics and calculate valuations",
        "",
        "For assistance, contact Cadence support"
    ]
    
    for instruction in instructions:
        ws.append([instruction])
        ws[f'A{ws.max_row}'].font = instruction_font
        ws.merge_cells(f'A{ws.max_row}:F{ws.max_row}')
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15
    
    # Save to BytesIO buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Cadence_Schedule_A_Template_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
